import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging

logger = logging.getLogger(__name__)


def generate_rev_and_inflow_report_sheet(workbook):
            # Create workbook and sheet
        try:
            sheet = workbook.create_sheet("Rev & Inflow")

            # Define styles
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            light_green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Light green
            light_blue_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # Light blue
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            sheet.column_dimensions['A'].width = 5
            sheet.column_dimensions['B'].width = 40
            sheet.column_dimensions['C'].width = 18
            sheet.column_dimensions['D'].width = 18
            sheet.column_dimensions['E'].width = 12
            for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
                sheet.column_dimensions[col].width = 15

            # Row 1: Summary totals header
            sheet.merge_cells('F1:G1')
            sheet.merge_cells('H1:H1')
            sheet.merge_cells('I1:I1')
            sheet.merge_cells('J1:J1')
            sheet.merge_cells('K1:K1')
            sheet.merge_cells('L1:L1')
            sheet.merge_cells('M1:M1')
            sheet.merge_cells('N1:N1')
            
            # Add placeholder values in row 1 (these would be dynamic totals)
            totals_row1 = [
                ('F1', '18,205,915'), ('H1', '-'), ('I1', '124,609,375'), ('J1', '469,712,613'),
                ('K1', '574,902,346'), ('L1', '639,634,489'), ('M1', '1,787,674,131'), ('N1', '2,244,040,884')
            ]
            for cell_ref, value in totals_row1:
                cell = sheet[cell_ref]
                cell.value = value
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[1].height = 25

            # Row 2: Main headers
            # SN and totals
            cell = sheet['A2']
            cell.value = 'SN'
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.alignment = center_alignment
            cell.border = border

            sheet.merge_cells('B2:E2')
            cell = sheet['B2']
            cell.value = '8,069,670,784   32.28'
            cell.font = bold_font
            cell.fill = light_green_fill
            cell.alignment = center_alignment
            cell.border = border

            # Revenue header spanning remaining columns
            sheet.merge_cells('F2:N2')
            cell = sheet['F2']
            cell.value = 'Revenue'
            cell.font = Font(bold=True, size=12)
            cell.alignment = center_alignment
            cell.border = border
            sheet.row_dimensions[2].height = 25

            # Row 3: Month headers
            cell = sheet['A3']
            cell.value = 'A'
            cell.font = bold_font
            cell.fill = light_green_fill
            cell.alignment = center_alignment
            cell.border = border

            cell = sheet['B3']
            cell.value = 'Contract-1 (Supply)'
            cell.font = bold_font
            cell.fill = light_green_fill
            cell.alignment = left_alignment
            cell.border = border

            # Price headers
            headers_row3 = [
                ('C3', 'Price for Supply\nFor 300MW'),
                ('D3', 'Price With GST'),
                ('E3', '5,656,375,577 Rs/Wp'),
                ('F3', '1\nJun-25'),
                ('G3', '2\nJul-25'),
                ('H3', '3\nAug-25'),
                ('I3', '4\nSep-25'),
                ('J3', '5\nOct-25'),
                ('K3', '6\nNov-25'),
                ('L3', '7\nDec-25'),
                ('M3', '8\nJan-26'),
                ('N3', '9\nFeb-26'),
            ]

            for cell_ref, text in headers_row3:
                cell = sheet[cell_ref]
                cell.value = text
                cell.font = bold_font
                if cell_ref in ['C3', 'D3', 'E3']:
                    cell.fill = light_green_fill
                else:
                    cell.fill = light_blue_fill
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[3].height = 35

            # Row 4 onwards: Item list headers (without data)
            items_supply = [
                'Solar Photovoltaic (SPV) Modules',
                'Mounting Structures (MMS)',
                'String Monitoring Unit (SMU)',
                'Power Conditioning Unit (PCU)',
                'Solar Cables (SPV Modules to SMU)',
                'DC Power Cables (SMU to Inverters)',
                'Inverter Transformers',
                'Auxiliary Power Transformer(s)',
                'High Voltage (HV) Switchgear',
                'LT Switchgear',
                'AC Cables (HT & LT)',
                'Battery System',
                'SCADA System',
                'Communication Cables',
                'Weather Monitoring Station',
                'Fire Fighting System',
                'Module Cleaning System (MCS)',
                'Earthing',
                'Balance of Systems including all the requisite equipment, materials, spares, accessories etc. excluding the items mentioned at Sr. No. 1.1 to 1.18 above.'
            ]

            current_row = 4
            for idx, item in enumerate(items_supply, start=1):
                # SN
                cell = sheet.cell(row=current_row, column=1)
                cell.value = idx
                cell.alignment = center_alignment
                cell.border = border

                # Item name
                cell = sheet.cell(row=current_row, column=2)
                cell.value = item
                cell.alignment = left_alignment
                cell.border = border

                # Empty cells for Price for Supply, Price With GST, Rs/Wp
                for col in [3, 4, 5]:
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = ''
                    cell.alignment = center_alignment
                    cell.border = border

                # Empty cells for months (F to N)
                for col in range(6, 15):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = ''
                    cell.alignment = center_alignment
                    cell.border = border

                sheet.row_dimensions[current_row].height = 25
                current_row += 1

            # Add Contract-2 (Service) section
            current_row += 1

            # Contract-2 header row
            cell = sheet[f'A{current_row}']
            cell.value = 'B'
            cell.font = bold_font
            cell.fill = light_blue_fill
            cell.alignment = center_alignment
            cell.border = border

            cell = sheet[f'B{current_row}']
            cell.value = 'Contract-2 (Service)'
            cell.font = bold_font
            cell.fill = light_blue_fill
            cell.alignment = left_alignment
            cell.border = border

            # Price headers for Contract-2
            contract2_headers = [
                (f'C{current_row}', 'Price for service'),
                (f'D{current_row}', '1,888,964,849'),
                (f'E{current_row}', '8'),
            ]

            for cell_ref, text in contract2_headers:
                cell = sheet[cell_ref]
                cell.value = text
                cell.font = bold_font
                cell.fill = light_blue_fill
                cell.alignment = center_alignment
                cell.border = border

            # Empty cells for months
            for col in range(6, 15):
                cell = sheet.cell(row=current_row, column=col)
                cell.value = ''
                cell.fill = light_blue_fill
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[current_row].height = 25
            current_row += 1

            # Service items
            items_service = [
                'Solar Photovoltaic (SPV) Modules',
                'Module Mounting Structures (MMS)',
                'String Monitoring Unit (SMU)',
                'Power Conditioning Unit (PCU)',
                'Solar Cables (SPV Modules to SMU)',
                'DC Power Cables (SMU to Inverters)',
                'Inverter Transformers',
                'Auxiliary Power Transformer(s)',
                'High Voltage (HV) Switchgear',
                'LT Switchgear',
                'AC Cables (HT & LT)',
                'Battery System',
                'SCADA System',
                'Communication Cables',
            ]

            for idx, item in enumerate(items_service, start=1):
                # SN
                cell = sheet.cell(row=current_row, column=1)
                cell.value = idx
                cell.alignment = center_alignment
                cell.border = border

                # Item name
                cell = sheet.cell(row=current_row, column=2)
                cell.value = item
                cell.alignment = left_alignment
                cell.border = border

                # Empty cells for prices and months
                for col in range(3, 15):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = ''
                    cell.alignment = center_alignment
                    cell.border = border

                sheet.row_dimensions[current_row].height = 25
                current_row += 1

            # Freeze panes at row 4, column F
            sheet.freeze_panes = 'F4'

          

        except Exception as e:
            logger.error(f"Error generating Payment Terms Outflow report: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)