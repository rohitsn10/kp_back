import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging

logger = logging.getLogger(__name__)


def generate_payment_terms_sheet(workbook):
        try:
            # Create workbook and sheet
            sheet = workbook.create_sheet("Payment Terms")

            # Define styles
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green
            light_blue_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # Light blue
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)
            white_bold_font = Font(bold=True, size=11, color="FFFFFF")

            # Set column widths
            sheet.column_dimensions['A'].width = 5
            sheet.column_dimensions['B'].width = 35
            sheet.column_dimensions['C'].width = 120
            sheet.column_dimensions['D'].width = 12
            sheet.column_dimensions['E'].width = 20

            # Row 2: Main title
            sheet.merge_cells('B2:E2')
            cell = sheet['B2']
            cell.value = 'Payment terms'
            cell.font = white_bold_font
            cell.fill = green_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.row_dimensions[2].height = 25

            # Row 3: Headers
            sheet.merge_cells('B3:C3')
            cell = sheet['B3']
            cell.value = '3.23 Cr/MWp'
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = border

            cell = sheet['D3']
            cell.value = 'Total Contract Price (In Rs)'
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = border

            cell = sheet['E3']
            cell.value = '8,069,670,784'
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = border

            sheet.row_dimensions[3].height = 20

            # Contract-1 (Supply) section
            current_row = 4

            # Contract-1 header
            cell = sheet[f'B{current_row}']
            cell.value = 'Contract-1 (Supply)'
            cell.font = bold_font
            cell.alignment = left_alignment
            cell.border = border

            cell = sheet[f'C{current_row}']
            cell.value = 'Schedule of Price for Supply of Plant and Equipment at site complete in all respect'
            cell.font = bold_font
            cell.alignment = left_alignment
            cell.border = border

            cell = sheet[f'D{current_row}']
            cell.value = '100%'
            cell.alignment = center_alignment
            cell.border = border

            cell = sheet[f'E{current_row}']
            cell.value = '5,656,375,577'
            cell.alignment = center_alignment
            cell.border = border

            sheet.row_dimensions[current_row].height = 20
            current_row += 1

            # Payment terms data for Contract-1
            contract1_terms = [
                {
                    'title': 'Advance',
                    'description': '''Paid upon submission of:
1. Unconditional acceptance of LOA and contract signing.
2. Bank Guarantee (110% of advance amount, valid till commissioning).
3. CPSG Bank Guarantee.
4. Detailed PERT network chart.
5. Alternatively, if the contractor opts out of advance payment, the 15% is added to dispatch payment, increasing it to 70%.''',
                    'percentage': '15%',
                    'amount': '848456336.6'
                },
                {
                    'title': 'On Dispatch',
                    'description': 'Pro-rata basis with 100% GST upon:\nSubmission of GST invoice, MDCC, packing list, indemnity bond, insurance policy, warranty certificate, local content certification, proof of dispatch, and a compliance certificate.',
                    'percentage': '55%',
                    'amount': '3111006568'
                },
                {
                    'title': 'On-Site Receipt Payment',
                    'description': 'Pro-rata basis upon submission of GST invoice, delivery acknowledgment, and physical verification report.',
                    'percentage': '10%',
                    'amount': '565637557.7'
                },
                {
                    'title': 'Post-Erection Payment',
                    'description': 'On erection completion certified by EIC.',
                    'percentage': '2.50%',
                    'amount': '141409389.4'
                },
                {
                    'title': 'Commissioning Payment',
                    'description': 'On issuing the commissioning certificate for the plant or part thereof.',
                    'percentage': '2.50%',
                    'amount': '141409389.4'
                },
                {
                    'title': 'Final Acceptance Payment',
                    'description': 'On final plant acceptance and work completion certificate.\nCan be released earlier post-commissioning against equivalent BG.',
                    'percentage': '15%',
                    'amount': '848456336.6'
                },
            ]

            for term in contract1_terms:
                cell = sheet[f'B{current_row}']
                cell.value = term['title']
                cell.alignment = left_alignment
                cell.border = border

                cell = sheet[f'C{current_row}']
                cell.value = term['description']
                cell.alignment = left_alignment
                cell.border = border

                cell = sheet[f'D{current_row}']
                cell.value = term['percentage']
                cell.alignment = center_alignment
                cell.border = border

                cell = sheet[f'E{current_row}']
                cell.value = term['amount']
                cell.alignment = center_alignment
                cell.border = border

                sheet.row_dimensions[current_row].height = 60
                current_row += 1

            # Empty row
            current_row += 1

            # Contract-2 Part-1 (Service) section
            cell = sheet[f'B{current_row}']
            cell.value = 'Contract-2 Part-1 (Service)'
            cell.font = bold_font
            cell.fill = light_blue_fill
            cell.alignment = left_alignment
            cell.border = border

            cell = sheet[f'C{current_row}']
            cell.value = 'Erection, Testing, Commissioning of Plant & Equipment, Performance Demonstration and Operational Acceptance including, Unloading, Handling at Site, Insurance Covers, Storage of the Plant & Equipment supplied under First Contract and all Civil, Architectural & Structural Works complete in all respect'
            cell.font = bold_font
            cell.fill = light_blue_fill
            cell.alignment = left_alignment
            cell.border = border

            cell = sheet[f'D{current_row}']
            cell.value = '100%'
            cell.fill = light_blue_fill
            cell.alignment = center_alignment
            cell.border = border

            cell = sheet[f'E{current_row}']
            cell.value = '1,888,964,849'
            cell.fill = light_blue_fill
            cell.alignment = center_alignment
            cell.border = border

            sheet.row_dimensions[current_row].height = 60
            current_row += 1

            # Payment terms for Contract-2 Part-1
            contract2_part1_terms = [
                {
                    'title': 'Mobilization Advance',
                    'description': '''Paid with interest (SBI MCLR + 50 basis points, compounded quarterly).
Recovery starts after 10% of contract payment and completes when 60% is paid or within 12 months.
Advance released upon submission of:
BG (110% of advance amount).
CPSG Bank Guarantee.''',
                    'percentage': '10%',
                    'amount': '188896484.9'
                },
                {
                    'title': 'Erection and Testing',
                    'description': 'Pro-rata basis with 100% GST upon:\nGST invoice, EIC certification, insurance proof, and local content certification.',
                    'percentage': '75%',
                    'amount': '1416723637'
                },
                {
                    'title': 'Commissioning',
                    'description': 'Paid on successful plant commissioning with EIC certification.',
                    'percentage': '10%',
                    'amount': '188896484.9'
                },
                {
                    'title': 'Final Acceptance',
                    'description': 'Paid on final acceptance with a local content certificate and work completion.',
                    'percentage': '5%',
                    'amount': '94448242.43'
                },
            ]

            for term in contract2_part1_terms:
                cell = sheet[f'B{current_row}']
                cell.value = term['title']
                cell.alignment = left_alignment
                cell.border = border

                cell = sheet[f'C{current_row}']
                cell.value = term['description']
                cell.alignment = left_alignment
                cell.border = border

                cell = sheet[f'D{current_row}']
                cell.value = term['percentage']
                cell.alignment = center_alignment
                cell.border = border

                cell = sheet[f'E{current_row}']
                cell.value = term['amount']
                cell.alignment = center_alignment
                cell.border = border

                sheet.row_dimensions[current_row].height = 60
                current_row += 1

            # Contract-2 Part-2 (Civil)
            cell = sheet[f'B{current_row}']
            cell.value = 'Contract-2 Part-2 (Civil)'
            cell.font = bold_font
            cell.fill = light_blue_fill
            cell.alignment = left_alignment
            cell.border = border

            cell = sheet[f'C{current_row}']
            cell.value = 'Civil Works under Second Contract'
            cell.font = bold_font
            cell.fill = light_blue_fill
            cell.alignment = left_alignment
            cell.border = border

            cell = sheet[f'D{current_row}']
            cell.value = '100%'
            cell.fill = light_blue_fill
            cell.alignment = center_alignment
            cell.border = border

            cell = sheet[f'E{current_row}']
            cell.value = '182,059,153'
            cell.fill = light_blue_fill
            cell.alignment = center_alignment
            cell.border = border

            sheet.row_dimensions[current_row].height = 20
            current_row += 1

            # Payment terms for Contract-2 Part-2
            contract2_part2_terms = [
                {
                    'title': 'Progressive Payment',
                    'description': 'Paid with 100% GST upon EIC certification of work milestones and quality assurance.',
                    'percentage': '75%',
                    'amount': '136544364.5'
                },
                {
                    'title': 'Commissioning',
                    'description': 'Paid on successful plant commissioning with EIC certification.',
                    'percentage': '10%',
                    'amount': '18205915.27'
                },
                {
                    'title': 'Final Acceptance',
                    'description': 'Paid on final acceptance with local content certification.',
                    'percentage': '15%',
                    'amount': '27308872.9'
                },
            ]

            for term in contract2_part2_terms:
                cell = sheet[f'B{current_row}']
                cell.value = term['title']
                cell.alignment = left_alignment
                cell.border = border

                cell = sheet[f'C{current_row}']
                cell.value = term['description']
                cell.alignment = left_alignment
                cell.border = border

                cell = sheet[f'D{current_row}']
                cell.value = term['percentage']
                cell.alignment = center_alignment
                cell.border = border

                cell = sheet[f'E{current_row}']
                cell.value = term['amount']
                cell.alignment = center_alignment
                cell.border = border

                sheet.row_dimensions[current_row].height = 30
                current_row += 1

            # Empty row
            current_row += 1

            # Contract-3 O&M
            cell = sheet[f'B{current_row}']
            cell.value = 'Contract-3- O&M'
            cell.font = bold_font
            cell.fill = light_blue_fill
            cell.alignment = left_alignment
            cell.border = border

            cell = sheet[f'C{current_row}']
            cell.value = '3 years from the Commercial Operation Date (COD) including O&M spares and consumables'
            cell.font = bold_font
            cell.fill = light_blue_fill
            cell.alignment = left_alignment
            cell.border = border

            cell = sheet[f'D{current_row}']
            cell.value = '100%'
            cell.fill = light_blue_fill
            cell.alignment = center_alignment
            cell.border = border

            cell = sheet[f'E{current_row}']
            cell.value = '342,271,206'
            cell.fill = light_blue_fill
            cell.alignment = center_alignment
            cell.border = border

            sheet.row_dimensions[current_row].height = 30
            current_row += 1

            # O&M payment terms description
            cell = sheet[f'C{current_row}']
            cell.value = '''1.The payment for Third Contract shall be made on quarterly basis including GST
2. The quarter will be defined as a period of three months ending on 30th June, 30th September, 31stDecember and 31st March except last quarter of the third contract wherein, payment will be made for the number of days covered till the last'''
            cell.alignment = left_alignment
            cell.border = border
            sheet.row_dimensions[current_row].height = 60

            # Save to response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="Payment_Terms.xlsx"'
            workbook.save(response)

            return response

        except Exception as e:
            logger.error(f"Error generating Payment Terms sheet: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)