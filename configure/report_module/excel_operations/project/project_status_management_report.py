import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from rest_framework.response import Response
import logging
import os
from django.conf import settings
from datetime import datetime

logger = logging.getLogger(__name__)


def generate_project_status_management_report(request):
    try:
        pass
        # Create workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Project Status Management"


        # Define styles
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Blue
        pink_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # Pink
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold_font = Font(bold=True, size=11)

        # Set column widths
        column_widths = {
            'A': 5,   # SN
            'B': 10,  # SAP Code
            'C': 15,  # Entity
            'D': 18,  # Project Category
            'E': 15,  # Client Name
            'F': 12,  # Location
            'G': 8,   # MWp
            'H': 8,   # MWac
            'I': 8,   # MMS
            'J': 12,  # Land lease
            'K': 15,  # Land Handover
            'L': 15,  # GEDA Provisional Approval
            'M': 12,  # TFR Approval
            'N': 15,  # Line Estimate Payment
            'O': 15,  # GEDA Final Approval
            'P': 12,  # CEI Plan Approval
            'Q': 15,  # CEI Charging Approval
            'R': 15,  # Wheeling Approval
            'S': 15,  # GEDA COD Approval
            'T': 10,  # MMS
            'U': 12,  # PV Modules
            'V': 12,  # String Inverters
            'W': 15,  # IDT Transformer
            'X': 10,  # ICOG
        }

        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        # Row 1: Section header "Statutory" (merged N1:S1)
        sheet.merge_cells('J1:S1')
        cell = sheet['J1']
        cell.value = 'Statutory'
        cell.font = bold_font
        cell.fill = yellow_fill
        cell.alignment = center_alignment
        cell.border = border

        # Row 1: Section header "Material Tracking" (merged T1:AA1)
        sheet.merge_cells('T1:AA1')
        cell = sheet['T1']
        cell.value = 'Material Tracking'
        cell.font = bold_font
        cell.fill = blue_fill
        cell.alignment = center_alignment
        cell.border = border

        # Row 1: Section header "Execution" (merged AB1:AQ1)
        sheet.merge_cells('AB1:AQ1')
        cell = sheet['AB1']
        cell.value = 'Execution'
        cell.font = bold_font
        cell.fill = pink_fill
        cell.alignment = center_alignment
        cell.border = border

        sheet.row_dimensions[1].height = 25

        # Row 2: Column headers
        headers = [
            ('A2', 'SN'),
            ('B2', 'SAP Code'),
            ('C2', 'Entity'),
            ('D2', 'Project Category'),
            ('E2', 'Client Name'),
            ('F2', 'Location'),
            ('G2', 'MWp'),
            ('H2', 'MWac'),
            ('I2', 'MMS'),
            ('J2', 'Land lease'),
            ('K2', 'Land Handover'),
            ('L2', 'GEDA Provisional\nApproval'),
            ('M2', 'TFR\nApproval'),
            ('N2', 'Line Estimate\nPayment'),
            ('O2', 'GEDA Final\nApproval'),
            ('P2', 'CEI Plan\nApproval'),
            ('Q2', 'CEI Charging\nApproval'),
            ('R2', 'Wheeling\nApproval'),
            ('S2', 'GEDA COD\nApproval'),
            ('T2', 'MMS'),
            ('U2', 'PV Modules'),
            ('V2', 'String\nInverters'),
            ('W2', 'IDT\nTransformer'),
            ('X2', 'ICOG'),
            ('Y2', 'LTDB'),  # Empty headers for alignment
            ('Z2', 'ABT & CTPT Unit (Plant End)'),
            ('AA2', 'ABT & CTPT Unit (GSS End)'),  
            ('AB2','Boundary Fencing'),
            ('AC2','Civil Piling'),
            ('AD2', 'Yard Civil Foundation'),
            ('AE2', 'Inverter Frame Foundation'),
            ('AF2', 'MMS Erection'),
            ('AG2','Module Installation'),
            ('AH2','Earthing Trench Excavation and Strip Laying'),
            ('AI2','AC Cable Trench Excavation and Cable Laying'),
            ('AJ2','DC Cable Trench Excavation and Cable Laying'),
            ('AK2','Communication cable laying'),
            ('AL2','Inverter Commissioning'),
            ('AM2','Yard Equipment Commissoning'),
            ('AN2','ABT & CTPT Sealing & Commissioning'),
            ('AO2','HDD Work'),
            ('AP2','HT Cable laying and Termination'),
            ('AQ2','VCB Integration @ GSS')

        ]

        for cell_ref, header_text in headers:
            cell = sheet[cell_ref]
            cell.value = header_text
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = border
            
          
        sheet.row_dimensions[2].height = 40

        # Sample data rows
        sample_data = [
            {
                "sn": 1,
                "sap_code": "123",
                "entity": "KPI",
                "project_category": "C&I",
                "client_name": "A",
                "location": "Vilayat",
                "mwp": "5.000",
                "mwac": "3.900",
                "mms": "FT",
                "land_lease": "1/4/2025",
                "land_handover": "8/4/2025",
                "geda_provisional": "15/4/2025",
                "tfr_approval": "22/4/2025",
                "line_estimate": "29/4/2025",
                "geda_final": "14/5/2025",
                "cei_plan": "29/5/2025",
                "cei_charging": "30/5/2025",
                "wheeling": "31/5/2025",
                "geda_cod": "1/6/2025",
                "mms_material": "2/6/2025",
                "pv_modules": "3/6/2025",
                "string_inverters": "4/6/2025",
                "idt_transformer": "5/6/2025",
                "icog": "6/6/2025"
            },
            {
                "sn": 2,
                "sap_code": "456",
                "entity": "SUNDEROI",
                "project_category": "MSEP",
                "client_name": "B",
                "location": "Jalipa",
                "mwp": "3.000",
                "mwac": "2.400",
                "mms": "HSAT",
                "land_lease": "",
                "land_handover": "",
                "geda_provisional": "",
                "tfr_approval": "",
                "line_estimate": "",
                "geda_final": "",
                "cei_plan": "",
                "cei_charging": "",
                "wheeling": "",
                "geda_cod": "",
                "mms_material": "",
                "pv_modules": "",
                "string_inverters": "",
                "idt_transformer": "",
                "icog": ""
            },
            {
                "sn": 3,
                "sap_code": "789",
                "entity": "KPIGEPL",
                "project_category": "KUSUM",
                "client_name": "C",
                "location": "Nizar",
                "mwp": "2.000",
                "mwac": "0.800",
                "mms": "HSAT",
                "land_lease": "",
                "land_handover": "",
                "geda_provisional": "",
                "tfr_approval": "",
                "line_estimate": "",
                "geda_final": "",
                "cei_plan": "",
                "cei_charging": "",
                "wheeling": "",
                "geda_cod": "",
                "mms_material": "",
                "pv_modules": "",
                "string_inverters": "",
                "idt_transformer": "",
                "icog": ""
            },
            {"sn": 4},
            {"sn": 5},
            {"sn": 6},
            {"sn": 7},
        ]

        # Add data rows starting from row 3
        current_row = 3
        for data in sample_data:
            # SN
            sheet.cell(row=current_row, column=1).value = data.get("sn", "")
            sheet.cell(row=current_row, column=1).alignment = center_alignment
            sheet.cell(row=current_row, column=1).border = border

            # All other columns
            columns = [
                ('B', 'sap_code'), ('C', 'entity'), ('D', 'project_category'),
                ('E', 'client_name'), ('F', 'location'), ('G', 'mwp'),
                ('H', 'mwac'), ('I', 'mms'), ('J', 'land_lease'),
                ('K', 'land_handover'), ('L', 'geda_provisional'), ('M', 'tfr_approval'),
                ('N', 'line_estimate'), ('O', 'geda_final'), ('P', 'cei_plan'),
                ('Q', 'cei_charging'), ('R', 'wheeling'), ('S', 'geda_cod'),
                ('T', 'mms_material'), ('U', 'pv_modules'), ('V', 'string_inverters'),
                ('W', 'idt_transformer'), ('X', 'icog')
            ]

            for col_letter, field_name in columns:
                col_index = ord(col_letter) - ord('A') + 1
                cell = sheet.cell(row=current_row, column=col_index)
                cell.value = data.get(field_name, "")
                cell.alignment = center_alignment
                cell.border = border
                
         

            sheet.row_dimensions[current_row].height = 20
            current_row += 1

        # Freeze panes at row 3, column B
        sheet.freeze_panes = 'B3'


        # Define the file path to save the report
        reports_dir = os.path.join(settings.MEDIA_ROOT, "reports/project/")
        os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
        datetime_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(reports_dir, f"project_status_management_report_{datetime_str}.xlsx")

        # Save the workbook to the file path
        workbook.save(file_path)

        # Construct the file URL
        file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/project/project_status_management_report_{datetime_str}.xlsx")
        return Response({"status": True, "message": "Project Status Management report generated successfully", "file_url": file_url}, status=200)

    except Exception as e:
        logger.error(f"Error generating report: {e}", exc_info=True)
        return Response({"error": str(e)}, status=500)