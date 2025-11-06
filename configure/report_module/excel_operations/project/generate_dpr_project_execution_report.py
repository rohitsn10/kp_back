import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging
import os
from django.conf import settings
from datetime import datetime

logger = logging.getLogger(__name__)


def generate_dpr_project_execution_report(request):
        try:
            # Create workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "DPR-Project Execution"

            # Define styles
            header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # Light blue
            section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Lighter blue
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            sheet.column_dimensions['A'].width = 5   # SN
            sheet.column_dimensions['B'].width = 35  # Activity Name
            sheet.column_dimensions['C'].width = 12  # Status
            sheet.column_dimensions['D'].width = 18  # Schedule Start Date
            sheet.column_dimensions['E'].width = 15  # Target Date
            sheet.column_dimensions['F'].width = 18  # Actual Start Date
            sheet.column_dimensions['G'].width = 18  # Completion Date

            # Main Header Row 1
            sheet.merge_cells('A1:G1')
            cell = sheet['A1']
            cell.value = 'Project Progress Report'
            cell.font = Font(bold=True, size=14)
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.row_dimensions[1].height = 25

            # Define sections and their activities
            sections = [
                {
                    "name": "Statutory Approvals",
                    "activities": [
                        "Land Lease",
                        "GEDA Provisional Approval",
                        "TFR Approval",
                        "GEDA Final Approval",
                        "CEI Plan Approval",
                        "CEI Charging Approval",
                        "Wheeling Approval",
                        "GEDA COD Approval"
                    ]
                },
                {
                    "name": "Major Material Delivery",
                    "activities": [
                        "MMS",
                        "PV Modules",
                        "String Inverters",
                        "IDT Transformer",
                        "ICOG",
                        "LTDB",
                        "Auxillary Transformer",
                        "ABT & CTPT Unit (Plant End)"
                    ]
                },
                {
                    "name": "Civil Activities",
                    "activities": [
                        "Boundary Fencing",
                        "Civil Piling",
                        "Yard Civil Foundation",
                        "Inverter Frame Foundation"
                    ]
                },
                {
                    "name": "Mechanical Activities",
                    "activities": [
                        "MMS Erection",
                        "Module Installation"
                    ]
                },
                {
                    "name": "Electrical Activities",
                    "activities": [
                        "Earthing Trench Excavation and Strip Laying",
                        "AC Cable Trench Excavation and Cable Laying",
                        "DC Cable Trench Excavation and Cable Laying",
                        "Communication cable laying",
                        "Inverter Commissioning",
                        "Yard Equipment Commissioning",
                        "ABT & CTPT Sealing & Commissioning"
                    ]
                },
                {
                    "name": "HT Line activities",
                    "activities": [
                        "HDD Work",
                        "HT Cable laying and Termination",
                        "VCB Integration @ GSS"
                    ]
                }
            ]

            current_row = 2

            for section in sections:
                # Section header row with SN label
                sheet.cell(row=current_row, column=1).value = "SN"
                sheet.cell(row=current_row, column=1).font = bold_font
                sheet.cell(row=current_row, column=1).fill = section_fill
                sheet.cell(row=current_row, column=1).alignment = center_alignment
                sheet.cell(row=current_row, column=1).border = border

                # Section name
                sheet.cell(row=current_row, column=2).value = section["name"]
                sheet.cell(row=current_row, column=2).font = bold_font
                sheet.cell(row=current_row, column=2).fill = section_fill
                sheet.cell(row=current_row, column=2).alignment = left_alignment
                sheet.cell(row=current_row, column=2).border = border

                # Column headers for this section
                headers = ["Status", "Schedule Start Date", "Target Date", "Actual Start Date", "Completion Date"]
                for idx, header in enumerate(headers, start=3):
                    cell = sheet.cell(row=current_row, column=idx)
                    cell.value = header
                    cell.font = bold_font
                    cell.fill = section_fill
                    cell.alignment = center_alignment
                    cell.border = border

                sheet.row_dimensions[current_row].height = 30
                current_row += 1

                # Activity rows
                for idx, activity in enumerate(section["activities"], start=1):
                    # Serial number
                    sheet.cell(row=current_row, column=1).value = idx
                    sheet.cell(row=current_row, column=1).alignment = center_alignment
                    sheet.cell(row=current_row, column=1).border = border

                    # Activity name
                    sheet.cell(row=current_row, column=2).value = activity
                    sheet.cell(row=current_row, column=2).alignment = left_alignment
                    sheet.cell(row=current_row, column=2).border = border

                    # Empty cells for data entry (Status, dates, etc.)
                    for col in range(3, 8):
                        cell = sheet.cell(row=current_row, column=col)
                        cell.value = ""
                        cell.alignment = center_alignment
                        cell.border = border

                    sheet.row_dimensions[current_row].height = 20
                    current_row += 1

            # Freeze panes at row 3, column B
            sheet.freeze_panes = 'B3'

            # Define the file path to save the report
            reports_dir = os.path.join(settings.MEDIA_ROOT, "reports/project/")
            os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
            datetime_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(reports_dir, f"dpr_project_execution_report_{datetime_str}.xlsx")

            # Save the workbook to the file path
            workbook.save(file_path)

            # Construct the file URL
            file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/project/dpr_project_execution_report_{datetime_str}.xlsx")
            return Response({"status": True, "message": "DPR Project Execution report generated successfully", "file_url": file_url}, status=200)


        except Exception as e:
            logger.error(f"Error generating DPR Project Execution report: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)