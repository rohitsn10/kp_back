import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging

logger = logging.getLogger(__name__)


class CashflowReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            # Create workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Summary"

            # Define styles
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            peach_fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")  # Peach/Light orange
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            sheet.column_dimensions['A'].width = 20
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
                sheet.column_dimensions[col].width = 10

            # Row 3: Title
            sheet.merge_cells('A3:P3')
            cell = sheet['A3']
            cell.value = 'Cashflow (In Cr.)'
            cell.font = Font(bold=True, size=14)
            cell.fill = peach_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.row_dimensions[3].height = 25

            # Row 4: Month headers
            months = [
                'Jun-25', 'Jul-25', 'Aug-25', 'Sep-25', 'Oct-25', 'Nov-25', 'Dec-25',
                'Jan-26', 'Feb-26', 'Mar-26', 'Apr-26', 'May-26', 'Jun-26', 'Jul-26', 'Aug-26'
            ]
            
            cell = sheet['A4']
            cell.value = 'Month'
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = border

            for idx, month in enumerate(months, start=2):
                cell = sheet.cell(row=4, column=idx)
                cell.value = month
                cell.font = bold_font
                cell.alignment = center_alignment
                cell.border = border
            
            sheet.row_dimensions[4].height = 20

            # Row 5: Revenue
            cell = sheet['A5']
            cell.value = 'Revenue'
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.alignment = left_alignment
            cell.border = border

            revenue_data = [1.8, '', '', 12.5, 47, 57, 64, 179, 224, 145, 40, 2, '', '', '']
            for idx, value in enumerate(revenue_data, start=2):
                cell = sheet.cell(row=5, column=idx)
                cell.value = value
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[5].height = 20

            # Rows 6: Inflow (Including 15% Advance) - merged with row 5 label area
            sheet.merge_cells('A6:A6')
            cell = sheet['A6']
            cell.value = 'Inflow\n(Including 15%\nAdvance)'
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.alignment = left_alignment
            cell.border = border

            inflow_data = [103.7, 1.82, '', 8, 6, 47, 34, 44, 119, 113, 97, 52, 32, 3, 11]
            for idx, value in enumerate(inflow_data, start=2):
                cell = sheet.cell(row=6, column=idx)
                cell.value = value
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[6].height = 45

            # Row 7: Outflow
            cell = sheet['A7']
            cell.value = 'Outflow'
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.alignment = left_alignment
            cell.border = border

            outflow_data = [0.7, 4, 38, 10, 9, 3, 24, 158, 240, 103, 6, 5, 6, '', '']
            for idx, value in enumerate(outflow_data, start=2):
                cell = sheet.cell(row=7, column=idx)
                cell.value = value
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[7].height = 20

            # Row 8: Net Cashflow
            cell = sheet['A8']
            cell.value = 'Net Cashflow'
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.alignment = left_alignment
            cell.border = border

            net_cashflow_data = [103.1, -2, -38, -2, -3, 44, 10, -115, -121, 10, 91, 47, 27, 3, 11]
            for idx, value in enumerate(net_cashflow_data, start=2):
                cell = sheet.cell(row=8, column=idx)
                cell.value = value
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[8].height = 20

            # Row 9: Absolute Cashflow
            cell = sheet['A9']
            cell.value = 'Absolute\nCashflow'
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.alignment = left_alignment
            cell.border = border

            absolute_cashflow_data = [103.1, 101, 63, 61, 58, 102, 111, -4, -124, -114, -23, 23, 50, 53, 16]
            for idx, value in enumerate(absolute_cashflow_data, start=2):
                cell = sheet.cell(row=9, column=idx)
                cell.value = value
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[9].height = 30

            # Freeze panes at row 4, column B
            sheet.freeze_panes = 'B4'

            # Save to response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="Cashflow_Report.xlsx"'
            workbook.save(response)

            return response

        except Exception as e:
            logger.error(f"Error generating Cashflow report: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)