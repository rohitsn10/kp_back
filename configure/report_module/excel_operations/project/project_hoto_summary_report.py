import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging
import os
from django.conf import settings
from datetime import datetime
logger = logging.getLogger(__name__)


def generate_hoto_summary_report(request):
        try:
            # Create workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Project Points Tracker"

            # Define styles
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Different colored fills for different sections
            electrical_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Light green
            civil_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # Light blue
            mms_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")  # Light purple
            document_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Light orange
            gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gray
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            column_widths = {
                'A': 8,   # SN
                'B': 20,  # Project Name
                'C': 12,  # HOTO Status
                'D': 15,  # Targetted Date
                'E': 18,  # EPC/Contractor Name
                'F': 12,  # Total Open Points
                'G': 12,  # Total Closed Points
                'H': 15,  # Total Balance Points
                'I': 10,  # ELECTRICAL Open
                'J': 10,  # ELECTRICAL Closed
                'K': 12,  # ELECTRICAL Balance
                'L': 10,  # CIVIL Open
                'M': 10,  # CIVIL Closed
                'N': 12,  # CIVIL Balance
                'O': 10,  # MMS Open
                'P': 10,  # MMS Closed
                'Q': 12,  # MMS Balance
                'R': 10,  # Document Open
                'S': 10,  # Document Closed
                'T': 12,  # Document Balance
                'U': 15,  # Remark
            }

            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Row 1: Section headers (merged cells)
            # ELECTRICAL points (I1:K1)
            sheet.merge_cells('I1:K1')
            cell = sheet['I1']
            cell.value = 'ELECTRICAL points'
            cell.font = bold_font
            cell.fill = electrical_fill
            cell.alignment = center_alignment
            cell.border = border

            # CIVIL points (L1:N1)
            sheet.merge_cells('L1:N1')
            cell = sheet['L1']
            cell.value = 'CIVIL points'
            cell.font = bold_font
            cell.fill = civil_fill
            cell.alignment = center_alignment
            cell.border = border

            # MMS points (O1:Q1)
            sheet.merge_cells('O1:Q1')
            cell = sheet['O1']
            cell.value = 'MMS points'
            cell.font = bold_font
            cell.fill = mms_fill
            cell.alignment = center_alignment
            cell.border = border

            # Document Submission status (R1:T1)
            sheet.merge_cells('R1:T1')
            cell = sheet['R1']
            cell.value = 'Document Submission status'
            cell.font = bold_font
            cell.fill = document_fill
            cell.alignment = center_alignment
            cell.border = border

            sheet.row_dimensions[1].height = 30

            # Row 2: Column headers
            headers = [
                ('A2', 'SN', gray_fill),
                ('B2', 'Project Name', gray_fill),
                ('C2', 'HOTO\nStatus', gray_fill),
                ('D2', 'Targetted\nDate', gray_fill),
                ('E2', 'EPC/Contractor\nName', gray_fill),
                ('F2', 'Total\nOpen\nPoints', gray_fill),
                ('G2', 'Total\nClosed\nPoints', gray_fill),
                ('H2', 'Total Balance\nPoints', gray_fill),
                ('I2', 'Open', electrical_fill),
                ('J2', 'Closed', electrical_fill),
                ('K2', 'Balance', electrical_fill),
                ('L2', 'Open', civil_fill),
                ('M2', 'Closed', civil_fill),
                ('N2', 'Balance', civil_fill),
                ('O2', 'Open', mms_fill),
                ('P2', 'Closed', mms_fill),
                ('Q2', 'Balance', mms_fill),
                ('R2', 'Open', document_fill),
                ('S2', 'Closed', document_fill),
                ('T2', 'Balance', document_fill),
                ('U2', 'Remark', gray_fill),
            ]

            for cell_ref, header_text, fill_color in headers:
                cell = sheet[cell_ref]
                cell.value = header_text
                cell.font = bold_font
                cell.fill = fill_color
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[2].height = 45

            # Sample data rows (you can replace with database data)
            sample_data = [
                {"sn": 1, "project_name": "", "hoto_status": "", "target_date": "", "epc_name": ""},
                {"sn": 2, "project_name": "", "hoto_status": "", "target_date": "", "epc_name": ""},
                {"sn": 3, "project_name": "", "hoto_status": "", "target_date": "", "epc_name": ""},
                {"sn": 4, "project_name": "", "hoto_status": "", "target_date": "", "epc_name": ""},
                {"sn": 5, "project_name": "", "hoto_status": "", "target_date": "", "epc_name": ""},
            ]

            # Add data rows starting from row 3
            current_row = 3
            for data in sample_data:
                # SN
                sheet.cell(row=current_row, column=1).value = data["sn"]
                sheet.cell(row=current_row, column=1).alignment = center_alignment
                sheet.cell(row=current_row, column=1).border = border

                # Project Name
                sheet.cell(row=current_row, column=2).value = data["project_name"]
                sheet.cell(row=current_row, column=2).alignment = center_alignment
                sheet.cell(row=current_row, column=2).border = border

                # HOTO Status
                sheet.cell(row=current_row, column=3).value = data["hoto_status"]
                sheet.cell(row=current_row, column=3).alignment = center_alignment
                sheet.cell(row=current_row, column=3).border = border

                # Targetted Date
                sheet.cell(row=current_row, column=4).value = data["target_date"]
                sheet.cell(row=current_row, column=4).alignment = center_alignment
                sheet.cell(row=current_row, column=4).border = border

                # EPC/Contractor Name
                sheet.cell(row=current_row, column=5).value = data["epc_name"]
                sheet.cell(row=current_row, column=5).alignment = center_alignment
                sheet.cell(row=current_row, column=5).border = border

                # Total points columns (F, G, H) with formula
                # Total Open Points (F) - placeholder
                sheet.cell(row=current_row, column=6).value = "-"
                sheet.cell(row=current_row, column=6).alignment = center_alignment
                sheet.cell(row=current_row, column=6).border = border

                # Total Closed Points (G) - placeholder
                sheet.cell(row=current_row, column=7).value = "-"
                sheet.cell(row=current_row, column=7).alignment = center_alignment
                sheet.cell(row=current_row, column=7).border = border

                # Total Balance Points (H) - formula (can add =F3/G3 if needed)
                sheet.cell(row=current_row, column=8).value = f"=F{current_row}/G{current_row}"
                sheet.cell(row=current_row, column=8).alignment = center_alignment
                sheet.cell(row=current_row, column=8).border = border

                # ELECTRICAL points (I, J, K) - all "-"
                for col in range(9, 12):
                    sheet.cell(row=current_row, column=col).value = "-"
                    sheet.cell(row=current_row, column=col).alignment = center_alignment
                    sheet.cell(row=current_row, column=col).border = border

                # CIVIL points (L, M, N) - all "-"
                for col in range(12, 15):
                    sheet.cell(row=current_row, column=col).value = "-"
                    sheet.cell(row=current_row, column=col).alignment = center_alignment
                    sheet.cell(row=current_row, column=col).border = border

                # MMS points (O, P, Q) - all "-"
                for col in range(15, 18):
                    sheet.cell(row=current_row, column=col).value = "-"
                    sheet.cell(row=current_row, column=col).alignment = center_alignment
                    sheet.cell(row=current_row, column=col).border = border

                # Document Submission status (R, S, T) - all "-"
                for col in range(18, 21):
                    sheet.cell(row=current_row, column=col).value = "-"
                    sheet.cell(row=current_row, column=col).alignment = center_alignment
                    sheet.cell(row=current_row, column=col).border = border

                # Remark (U)
                sheet.cell(row=current_row, column=21).value = ""
                sheet.cell(row=current_row, column=21).alignment = center_alignment
                sheet.cell(row=current_row, column=21).border = border

                sheet.row_dimensions[current_row].height = 25
                current_row += 1

            # Freeze panes at row 3, column B
            sheet.freeze_panes = 'B3'

            # Define the file path to save the report
            reports_dir = os.path.join(settings.MEDIA_ROOT, "reports/project/")
            os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
            datetime_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(reports_dir, f"hoto_summary_report_{datetime_str}.xlsx")

            # Save the workbook to the file path
            workbook.save(file_path)

            # Construct the file URL
            file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/project/hoto_summary_report_{datetime_str}.xlsx")
            return Response({"status": True, "message": "Hoto Summary report generated successfully", "file_url": file_url}, status=200)


        except Exception as e:
            logger.error(f"Error generating Project Points Tracker: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)