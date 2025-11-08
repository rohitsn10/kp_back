import openpyxl

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
from datetime import datetime
from django.conf import settings
from rest_framework.response import Response

def generate_iar_report(request):
    """
    Generate an Excel sheet matching the IAR document format.

    Args:
        filename (str): Output filename for the Excel file
    """
    try:
        # Create workbook and select active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Insurance Document"
        
        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        light_green_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        
        # Merge cells for company name
        ws.merge_cells('A1:J3')
        ws['A1'] = "AUM Insurance Brokers Pvt Ltd"
        ws['A1'].font = Font(size=18, bold=True, color="1F4E78")
        ws['A1'].alignment = Alignment(horizontal="right", vertical="center")
        
        # Name of Insured
        ws.merge_cells('A4:B4')
        ws['A4'] = "Name of Insured"
        ws['A4'].fill = header_fill
        ws['A4'].font = header_font
        ws['A4'].alignment = left_align
        
        ws.merge_cells('C4:J4')
        ws['C4'] = "Dhanpriya Prints Pvt. Ltd"
        ws['C4'].fill = light_green_fill
        ws['C4'].alignment = left_align
        
        # GSTIN
        ws.merge_cells('A5:B5')
        ws['A5'] = "GSTIN"
        ws['A5'].fill = header_fill
        ws['A5'].font = header_font
        ws['A5'].alignment = left_align
        
        ws.merge_cells('C5:J5')
        ws['C5'] = "24AABCD9653L1ZC"
        ws['C5'].fill = light_green_fill
        ws['C5'].alignment = left_align
        
        # Correspondence Address
        ws.merge_cells('A6:B6')
        ws['A6'] = "Correspondence Address"
        ws['A6'].fill = header_fill
        ws['A6'].font = header_font
        ws['A6'].alignment = left_align
        
        ws.merge_cells('C6:J6')
        ws['C6'] = "Plot No 13, Block No. 20-21, Behind Garden Vareli, Kadodara,, Palsana, SURAT, Gujarat - 394327"
        ws['C6'].alignment = left_align

        ws.merge_cells('C7:J7')
        ws['C7'] = ""

        ws.merge_cells('C9:J9')
        ws['C9'] = ""

        ws.merge_cells('C10:J10')
        ws['C10'] = ""

        ws.merge_cells('C11:J11')
        ws['C11'] = ""

        ws.merge_cells('C12:J12')
        ws['C12'] = ""

        ws.merge_cells('C13:J13')
        ws['C13'] = ""

        ws.merge_cells('C14:J14')
        ws['C14'] = ""

        ws.merge_cells('C15:J15')
        ws['C15'] = ""

        # Risk Occupancy
        ws.merge_cells('A7:B7')
        ws['A7'] = "Risk Occupancy"
        ws['A7'].fill = header_fill
        ws['A7'].font = header_font
        ws['A7'].alignment = left_align
        
        # Risk Location
        ws.merge_cells('A8:B8')
        ws['A8'] = "Risk Location"
        ws['A8'].fill = header_fill
        ws['A8'].font = header_font
        ws['A8'].alignment = left_align
        
        ws.merge_cells('C8:J8')
        ws['C8'] = "Survey No. 119, Vil: Namalpur, Ta: Tilakwada, Dist: Narmada"
        ws['C8'].alignment = left_align
        
        # Risk Code, Type of Policy, Policy Period, Previous Policy Details
        for row, label in [(9, "Risk Code (Tariff)"), (10, "Type of Policy"), 
                        (11, "Policy Period")]:
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = label
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].alignment = left_align

       # Write + style BEFORE merging
        ws['A12'] = "Previous Policy Details"
        ws['A12'].fill = header_fill
        ws['A12'].font = header_font
        ws['A12'].alignment = left_align

        # Now merge
        ws.merge_cells('A12:B13')

        # Claim History, Financial Interest
        for row, label in [(14, "Claim History"), (15, "Financial Interest (HP)")]:
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = label
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].alignment = left_align
        
        # Table headers
        headers = [
            ("A16", "Sr No"),
            ("B16", "Description of M/c"),
            ("C16", "Make/Model"),
            ("D16", "UOM"),
            ("E16", "Quantity"),
            ("F16", "Sr No"),
            ("G16", "Sum Insured")
        ]
        
        for cell, text in headers:
            ws[cell] = text
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = center_align
            ws[cell].border = thin_border
        
        # Equipment data
        equipment_data = [
            (1, "PV Module (545 Wp)", "WAAREE /Bi-SS-54S", "Nos", "", "", ""),
            (2, "String Inverter", "Sungrow/SG320HX", "Nos", "", "Attached in Sheet", ""),
            (3, "MMS", "KP standards", "Mw", "", "-", ""),
            (4, "Inverter Duty Transformer 2700kVA, 800-800/11000 Volts, 2W, 1HV Winding Transformer", "NAKODA PRODUCT", "Nos", "", "NP-23-148", ""),
            (5, "15kVA, 800/433V Auxiliary Transformer(For Common Infra)", "NAKODA PRODUCT", "Nos", "", "NP-23-150", ""),
            (6, "LTDB Panel- 7In 1Out, 2000A", "VINAYAK ELECTRICAL", "Nos", "", "VE/EP/73-1/23-24", ""),
            (7, "Auxiliary LTDB Panel (For Common Infra)", "ALNICO ELECTRIC", "Nos", "", "ALNC002060424/1", ""),
            (8, "11kV,800A Outgoing VCB without integrated adaptor cubical panels bearing(common Infra)", "VINAYAK ELECTRICAL(L&T)", "Nos", "", "M-2023-375", ""),
            (9, "11kV, 1250A Outgoing VCB with integrated adaptor cubical panels bearing((For Common Infra)", "JYOTI", "Nos", "", "73241947", ""),
            (10, "ABT Meter(plant End)", "SECURE", "Nos", "", "DG-0413B, DG-0414B", ""),
            (11, "11kV CTPT", "Varsha Engineers", "Set", "", "188/2023", ""),
            (12, "Other Items with No serial nos : (HT/LT/DC/OFC/RS485 Cables, HDPE Pipes, Civil Pilings, MMS, GI Structures, Equipment Foundations, Roads, Fencing, 11KV Lines, SCADA Network, I&C Labour, etc.", "As Per KP Standards", "Nos", "", "-", "")
        ]
        
        row_num = 17
        for data in equipment_data:
            sr_no, desc, make, uom, qty, sr_no2, sum_insured = data
            ws[f'A{row_num}'] = sr_no
            ws[f'B{row_num}'] = desc
            ws[f'C{row_num}'] = make
            ws[f'D{row_num}'] = uom
            ws[f'E{row_num}'] = qty
            ws[f'F{row_num}'] = sr_no2
            ws[f'G{row_num}'] = sum_insured
            
            # Apply borders and alignment
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                cell = ws[f'{col}{row_num}']
                cell.border = thin_border
                if col in ['A', 'D']:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            
            row_num += 1
        
        # Total row
        ws.merge_cells(f'A{row_num}:F{row_num}')
        ws[f'A{row_num}'] = "Total Value(with GST)"
        ws[f'A{row_num}'].fill = yellow_fill
        ws[f'A{row_num}'].font = Font(bold=True)
        ws[f'A{row_num}'].alignment = center_align
        ws[f'A{row_num}'].border = thin_border
        
        ws[f'G{row_num}'].border = thin_border
        
        # Define the file path to save the report
        reports_dir = os.path.join(settings.MEDIA_ROOT, "reports/project/")
        os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
        datetime_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(reports_dir, f"project_iar_report_{datetime_str}.xlsx")

        # Save the workbook to the file path
        wb.save(file_path)

        # Construct the file URL
        file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/project/project_iar_report_{datetime_str}.xlsx")
        return Response({"status": True, "message": "Project IAR report generated successfully", "file_url": file_url}, status=200)

    except Exception as e:
        print(f"Error generating Excel file: {e}")
        return Response({"status": False, "message": "Error generating report"}, status=500)

