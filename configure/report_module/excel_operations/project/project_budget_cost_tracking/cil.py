import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging

logger = logging.getLogger(__name__)


def generate_cil_report(request):
        try:
            # Create workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "CIL"

            # Define styles
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            column_widths = {
                'A': 15,  # Subproject No.
                'B': 20,  # SubProject Name
                'C': 30,  # Stage Name
                'D': 50,  # Description
                'E': 18,  # Planned Cost
                'F': 25,  # Invoiced Amount (A/R Open
                'G': 25,  # Amount (A/R Invoiced
                'H': 25,  # Amount (A/P) Open
                'I': 25,  # Amount (A/P)
                'J': 30,  # Project Name
            }

            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Row 1: Column headers
            headers = [
                ('A1', 'Subproject No.'),
                ('B1', 'SubProject Name'),
                ('C1', 'Stage Name'),
                ('D1', 'Description'),
                ('E1', 'Planned Cost'),
                ('F1', 'Invoiced Amount (A/R Open'),
                ('G1', 'Amount (A/R Invoiced'),
                ('H1', 'Amount (A/P) Open'),
                ('I1', 'Amount (A/P)'),
                ('J1', 'Project Name'),
                ('K1','BP Name'),
                ('L1','Total Planned Cost'),
                ('M1','Project Start Date'),
                ('N1','Due Date'),
                ('O1','Status'),
                ('P1','Internal Key'),
                ('Q1','Document Number'),
                ('R1','Series'),

            ]

            for cell_ref, header_text in headers:
                cell = sheet[cell_ref]
                cell.value = header_text
                cell.font = bold_font
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[1].height = 30

            # Sample data rows (empty values, ready for data entry)
            sample_data = [
                {
                    'subproject_no': '1',
                    'subproject_name': 'PV Modules',
                    'stage_name': 'Modules',
                    'description': 'PV Modules',
                    'planned_cost': '5,378,213,383.00',
                    'invoiced_ar_open': '0',
                    'amount_ar_invoiced': '0',
                    'amount_ap_open': '0',
                    'amount_ap': '5,224,500,000.00',
                    'project_name': 'Coal India Limited-S-4'
                },
                {
                    'subproject_no': '2',
                    'subproject_name': 'Inverters',
                    'stage_name': 'Inverter',
                    'description': 'Inverters',
                    'planned_cost': '358,547,559.00',
                    'invoiced_ar_open': '0',
                    'amount_ar_invoiced': '0',
                    'amount_ap_open': '0',
                    'amount_ap': '338,223,600.00',
                    'project_name': 'Coal India Limited-S-4'
                },
                {
                    'subproject_no': '3',
                    'subproject_name': 'MMS with Acc.',
                    'stage_name': 'MMS with Acc.',
                    'description': 'MMS with Acc.',
                    'planned_cost': '2,339,674,748.00',
                    'invoiced_ar_open': '0',
                    'amount_ar_invoiced': '0',
                    'amount_ap_open': '0',
                    'amount_ap': '590,812,245.30',
                    'project_name': 'Coal India Limited-S-4'
                },
                {
                    'subproject_no': '4',
                    'subproject_name': 'EPCC-BOS',
                    'stage_name': 'Balance of System',
                    'description': 'DC PACKAGE',
                    'planned_cost': '643,649,888.00',
                    'invoiced_ar_open': '0',
                    'amount_ar_invoiced': '0',
                    'amount_ap_open': '0',
                    'amount_ap': '0',
                    'project_name': 'Coal India Limited-S-4'
                },
                {
                    'subproject_no': '4',
                    'subproject_name': 'EPCC-BOS',
                    'stage_name': 'Balance of System',
                    'description': 'AC PACKAGE',
                    'planned_cost': '744,805,140.00',
                    'invoiced_ar_open': '0',
                    'amount_ar_invoiced': '0',
                    'amount_ap_open': '0',
                    'amount_ap': '0',
                    'project_name': 'Coal India Limited-S-4'
                },
                {
                    'subproject_no': '4',
                    'subproject_name': 'EPCC-BOS',
                    'stage_name': 'Balance of System',
                    'description': 'CIVIL & STRUCTURE',
                    'planned_cost': '672,737,410.00',
                    'invoiced_ar_open': '0',
                    'amount_ar_invoiced': '0',
                    'amount_ap_open': '0',
                    'amount_ap': '130,319,280.00',
                    'project_name': 'Coal India Limited-S-4'
                },
                {
                    'subproject_no': '5',
                    'subproject_name': 'Power evacuation(33 KV LINE',
                    'stage_name': 'Common 33kV Line Supply & services (UG/OHL',
                    'description': 'Common 33kV Line Supply & services (UG/OHL',
                    'planned_cost': '396,900,731.00',
                    'invoiced_ar_open': '0',
                    'amount_ar_invoiced': '0',
                    'amount_ap_open': '0',
                    'amount_ap': '0',
                    'project_name': 'Coal India Limited-S-4'
                },
                {
                    'subproject_no': '6',
                    'subproject_name': 'Others',
                    'stage_name': 'Contingency, Miscllaneous & Others',
                    'description': 'Others',
                    'planned_cost': '409,704,714.00',
                    'invoiced_ar_open': '0',
                    'amount_ar_invoiced': '0',
                    'amount_ap_open': '4,412,737.46',
                    'amount_ap': '17,008,902.14',
                    'project_name': 'Coal India Limited-S-4'
                },
            ]

            # Add data rows starting from row 2
            current_row = 2
            for data in sample_data:
                columns_data = [
                    (1, data['subproject_no'], center_alignment),
                    (2, data['subproject_name'], left_alignment),
                    (3, data['stage_name'], left_alignment),
                    (4, data['description'], left_alignment),
                    (5, data['planned_cost'], center_alignment),
                    (6, data['invoiced_ar_open'], center_alignment),
                    (7, data['amount_ar_invoiced'], center_alignment),
                    (8, data['amount_ap_open'], center_alignment),
                    (9, data['amount_ap'], center_alignment),
                    (10, data['project_name'], left_alignment),
                ]

                for col, value, alignment in columns_data:
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = value
                    cell.alignment = alignment
                    cell.border = border

                sheet.row_dimensions[current_row].height = 25
                current_row += 1

            # Add project name in row 12
            current_row = 12
            cell = sheet[f'A{current_row}']
            cell.value = 'coal india'
            cell.alignment = left_alignment

            # Add empty rows for additional data entry (rows 13-30)
            for row in range(13, 31):
                for col in range(1, 11):
                    cell = sheet.cell(row=row, column=col)
                    cell.value = ''
                    cell.alignment = center_alignment
                    cell.border = border
                sheet.row_dimensions[row].height = 20

            # Freeze panes at row 2, column A
            sheet.freeze_panes = 'A2'

            return Response({'workbook': workbook}, status=200)

        except Exception as e:
            logger.error(f"Error generating Budget Vs Actual tracker: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)