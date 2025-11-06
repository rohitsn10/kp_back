import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from rest_framework.response import Response
import logging
import os
from django.conf import settings
from datetime import datetime

logger = logging.getLogger(__name__)


def generate_ear_report(request):
    try:
        # Create workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "EAR Format"

        # Define styles
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        bold_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gray

        # Set column widths
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 60
        sheet.column_dimensions['C'].width = 50
        sheet.column_dimensions['D'].width = 20

        # Merge cells for logo (Row 2-7, Column B)
        sheet.merge_cells('B2:B7')
        cell = sheet['B2']
        cell.value = "KP LOGO"  # Placeholder for logo
        cell.alignment = center_alignment
        cell.border = border
        # You can add logo image here if needed
        # img = Image('path/to/kp_logo.png')
        # sheet.add_image(img, 'B2')

        # Company info (Row 2-7, Column C)
        sheet.merge_cells('C2:C7')
        company_info = "KP Group\nKP House, Opposite Ishwar farm junction BRTS, Near KP Circle,\nBhatar Canal road, Surat-395017"
        cell = sheet['C2']
        cell.value = company_info
        cell.alignment = center_alignment
        cell.border = border
        cell.font = Font(size=10)

        # Solarism logo area (Row 2-7, Column D)
        sheet.merge_cells('D2:D7')
        cell = sheet['D2']
        cell.value = "SOLARISM\nThe Power of Nature"
        cell.alignment = center_alignment
        cell.border = border
        cell.font = Font(bold=True, size=14, color="FF6600")

        # Information header (Row 8, spanning B to D)
        sheet.merge_cells('B8:D8')
        cell = sheet['B8']
        cell.value = "Information"
        cell.alignment = center_alignment
        cell.border = border
        cell.font = bold_font
        cell.fill = header_fill
        sheet.row_dimensions[8].height = 25

        # Row 9: Column headers
        # Project Name header
        cell = sheet['B9']
        cell.value = "Project Name"
        cell.alignment = left_alignment
        cell.border = border
        cell.font = bold_font

        # Answered header
        cell = sheet['C9']
        cell.value = "Answered"
        cell.alignment = center_alignment
        cell.border = border
        cell.font = bold_font

        # Merge D9 (if needed for alignment)
        sheet.merge_cells('D9:D9')
        cell = sheet['D9']
        cell.value = ""
        cell.border = border

        sheet.row_dimensions[9].height = 25

        # Project information data
        project_data = [
            {"sl": 1, "question": "Type of Project/Plant (Solar/Wind mills)", "answer": "Solar Power Plant"},
            {"sl": 2, "question": "Principle Name", "answer": "Ekta Prints Pvt Ltd"},
            {"sl": 3, "question": "Insured Name", "answer": "Sun Drops Energia Pvt Ltd"},
            {"sl": 4, "question": "Project Start Date", "answer": "1/4/2025"},
            {"sl": 5, "question": "Project COD", "answer": "30/9/2025"},
            {"sl": 6, "question": "Risk Location address (Village Name/City/District/State/Pin code", "answer": "Survey No. 502 Village-Dahej, Ta-Bharuch, Dist.\nBharuch-392130"},
            {"sl": 7, "question": "This is an extension of existing Plant or New Plant (Greenfield) ?", "answer": "New Plant"},
            {"sl": 8, "question": "Capacity of the New Plant", "answer": "1.5 MWp"},
            {"sl": 9, "question": "Capacity of the existing Plant (in case of extension project)", "answer": "-"},
            {"sl": 10, "question": "If the project is extension of existing plant, please provide distance between the existing plant & expansion project ?", "answer": "-"},
            {"sl": 11, "question": "Distance of Risk Location from Nearest Water body (canal/river/sea )", "answer": "8 Km  from Sea"},
            {"sl": 12, "question": "Whether integrated testing will be carried out with the existing plant ?", "answer": "NA"},
            {"sl": 13, "question": "Whether any used/second hand items/machinery will be erected/ tested ?", "answer": "NA"},
            {"sl": 14, "question": "If so. Please provide the details with Bifurcated Sum Insured", "answer": "-"},
            {"sl": 15, "question": "Availability of closed Storage facility ?", "answer": "-"},
            {"sl": 16, "question": "Security Guard/CCTV Surveillance arrangements ?", "answer": "Available"},
            {"sl": 17, "question": "Availability of Firefighting equipment at sites ?", "answer": "-"},
            {"sl": 18, "question": "Is there any Compound wall Available ?", "answer": "Barbed wire with chainlink fencing"},
            {"sl": 19, "question": "Nearest Fire Station Distance ?", "answer": "7 Km"},
            {"sl": 20, "question": "It is located at low line area (Flood/inundation experience any in past) ?", "answer": "No"},
        ]

        # Add project data rows starting from row 10
        current_row = 10
        for item in project_data:
            # Serial number (Column A)
            cell = sheet.cell(row=current_row, column=1)
            cell.value = item["sl"]
            cell.alignment = center_alignment
            cell.border = border

            # Question (Column B)
            cell = sheet.cell(row=current_row, column=2)
            cell.value = item["question"]
            cell.alignment = left_alignment
            cell.border = border
            if item.get("bold", False):
                cell.font = bold_font

            # Answer (Column C-D merged)
            sheet.merge_cells(f'C{current_row}:D{current_row}')
            cell = sheet.cell(row=current_row, column=3)
            cell.value = item["answer"]
            cell.alignment = left_alignment
            cell.border = border

            # Set row height based on content
            sheet.row_dimensions[current_row].height = 30 if '\n' in str(item.get("answer", "")) else 20
            current_row += 1

        # Financial section header (empty row for spacing if needed)
        sheet.merge_cells(f'B{current_row}:D{current_row}')
        cell = sheet.cell(row=current_row, column=2)
        cell.value = "SUM INSURED IN EPCC BOS WITH GST AMOUNT (Contractor Scope)"
        cell.alignment = left_alignment
        cell.border = border
        cell.font = bold_font
        sheet.row_dimensions[current_row].height = 25
        current_row += 1

        # Financial data
        financial_data = [
            {"item": "MODULES WITH GST AMOUNT (KP SCOPE)", "amount": 24360000},
            {"item": "INVERTERS WITH GST AMOUNT (KP SCOPE)", "amount": 2128000},
            {"item": "MMS BASIC WITH GST AMOUNT (KP SCOPE)", "amount": 3186000},
        ]

        # Add financial rows
        for item in financial_data:
            # Item description (Column B-C merged)
            sheet.merge_cells(f'B{current_row}:C{current_row}')
            cell = sheet.cell(row=current_row, column=2)
            cell.value = item["item"]
            cell.alignment = left_alignment
            cell.border = border

            # Amount (Column D)
            cell = sheet.cell(row=current_row, column=4)
            cell.value = item["amount"]
            cell.alignment = center_alignment
            cell.border = border
            cell.number_format = '#,##0'
            
            sheet.row_dimensions[current_row].height = 20
            current_row += 1

        # Total row
        sheet.merge_cells(f'B{current_row}:C{current_row}')
        cell = sheet.cell(row=current_row, column=2)
        cell.value = "Total Sum Insured (KP SCOPE)"
        cell.alignment = left_alignment
        cell.border = border
        cell.font = bold_font
        cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")

        cell = sheet.cell(row=current_row, column=4)
        total_amount = sum(item["amount"] for item in financial_data)
        cell.value = total_amount
        cell.alignment = center_alignment
        cell.border = border
        cell.font = bold_font
        cell.number_format = '#,##0'
        cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
        
        sheet.row_dimensions[current_row].height = 25

        # Define the file path to save the report
        reports_dir = os.path.join(settings.MEDIA_ROOT, "reports/project/")
        os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
        datetime_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(reports_dir, f"ear_report_{datetime_str}.xlsx")

        # Save the workbook to the file path
        workbook.save(file_path)

        # Construct the file URL
        file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/project/ear_report_{datetime_str}.xlsx")
        return Response({"status": True, "message": "EAR report generated successfully", "file_url": file_url}, status=200)

    except Exception as e:
        logger.error(f"Error generating EAR Format report: {e}", exc_info=True)
        return Response({"status": False, "error": str(e)}, status=500)