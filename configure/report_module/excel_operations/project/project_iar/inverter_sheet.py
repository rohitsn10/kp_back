from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def generate_inverter_sheet(workbook):
    """
    Generate the inverter sheet for the project payment terms report.
    
    Args:
        inverter_data (list): List of dictionaries containing inverter data.
    """

    try:
        # Create a new sheet for inverters
        inverter_sheet = workbook.create_sheet(title="Inverter")

        # Define styless
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set column widths
        columns = ['A', 'B']
        widths = [10, 30]
        for col, width in zip(columns, widths):
            inverter_sheet.column_dimensions[col].width = width

        # Header row
        headers = ["S.No", "Inverter Sr. No."]
        inverter_sheet.append(headers)

        # Apply header styles
        for cell in inverter_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Sample data for the inverter sheet
        inverter_data = [
            {"Sr. No.": 1, "Inverter Sr. No.": "A23B3001124"},
            {"Sr. No.": 2, "Inverter Sr. No.": "A23B3001060"},
            {"Sr. No.": 3, "Inverter Sr. No.": "A23B3001115"},
            {"Sr. No.": 4, "Inverter Sr. No.": "A23B3001128"},
            {"Sr. No.": 5, "Inverter Sr. No.": "A23C0100731"},
            {"Sr. No.": 6, "Inverter Sr. No.": "A23B3001109"},
            {"Sr. No.": 7, "Inverter Sr. No.": "A23C0100789"}
        ]

        # Populate data rows
        for row_data in inverter_data:
            row = [row_data["Sr. No."], row_data["Inverter Sr. No."]]
            inverter_sheet.append(row)

        # Apply styles to data rows
        for row in inverter_sheet.iter_rows(min_row=2, max_row=inverter_sheet.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border

    except Exception as e:
        print(f"Error generating inverter sheet: {e}")