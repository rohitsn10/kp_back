import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging
from django.conf import settings
import os
from datetime import datetime
logger = logging.getLogger(__name__)

def generate_project_delay_analysis_report(request):
        try:
            # Create workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Project Delay Analysis"

            # Define styles
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            column_widths = {
                'A': 8,   # SN
                'B': 25,  # Project Name
                'C': 18,  # Date of Welcome mail
                'D': 20,  # Committed date of Commissioning
                'E': 15,  # DC Capacity (MWp)
                'F': 15,  # AC Capacity (MWp)
                'G': 25,  # Delay of commissioning in days (Aging)
                'H': 25,  # Activities Delayed
                'I': 20,  # Revised Timelines - Revision 1
                'J': 20,  # Revised Timelines - Revision 2
                'K': 20,  # Revised Timelines - Revision 3
                'L': 35,  # Remarks/Justification of Delay
            }

            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Row 4: Section headers
            # Project Initial Details (C4:F4)
            sheet.merge_cells('C4:F4')
            cell = sheet['C4']
            cell.value = 'Project Initial Details'
            cell.font = bold_font
            cell.fill = white_fill
            cell.alignment = center_alignment
            cell.border = border

            # Aging Calculation (G4:G4)
            cell = sheet['G4']
            cell.value = 'Aging Calculation'
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.alignment = center_alignment
            cell.border = border

            sheet.row_dimensions[4].height = 25

            # Row 5: Column headers
            headers = [
                ('A5', 'SN', white_fill),
                ('B5', 'Project Name', white_fill),
                ('C5', 'Date of Welcome\nmail', white_fill),
                ('D5', 'Committed date of\nCommissioning', white_fill),
                ('E5', 'DC Capacity\n(MWp)', white_fill),
                ('F5', 'AC Capacity\n(MWp)', white_fill),
                ('G5', 'Delay of commissioning in\ndays (Aging)', yellow_fill),
                ('H5', 'Activities Delayed', white_fill),
                ('I5', 'Revised\nTimelines -\nRevision 1', white_fill),
                ('J5', 'Revised\nTimelines -\nRevision 2', white_fill),
                ('K5', 'Revised\nTimelines -\nRevision 3', white_fill),
                ('L5', 'Remarks/Justification of Delay', white_fill),
            ]

            for cell_ref, header_text, fill_color in headers:
                cell = sheet[cell_ref]
                cell.value = header_text
                cell.font = bold_font
                cell.fill = fill_color
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[5].height = 50

            # Sample data row
            sample_data = [
                {
                    "sn": 1,
                    "project_name": "Name of Client Customer",
                    "welcome_date": "10/4/2024",
                    "committed_date": "2/2/2025",
                    "dc_capacity": "5",
                    "ac_capacity": "4",
                    "delay_days": "278",
                    "activities_delayed": "Connectivity Approval",
                    "revision_1": "",
                    "revision_2": "",
                    "revision_3": "",
                    "remarks": ""
                }
            ]

            # Add data rows starting from row 6
            current_row = 6
            for data in sample_data:
                # SN
                sheet.cell(row=current_row, column=1).value = data.get("sn", "")
                sheet.cell(row=current_row, column=1).alignment = center_alignment
                sheet.cell(row=current_row, column=1).border = border

                # Project Name
                sheet.cell(row=current_row, column=2).value = data.get("project_name", "")
                sheet.cell(row=current_row, column=2).alignment = center_alignment
                sheet.cell(row=current_row, column=2).border = border

                # Date of Welcome mail
                sheet.cell(row=current_row, column=3).value = data.get("welcome_date", "")
                sheet.cell(row=current_row, column=3).alignment = center_alignment
                sheet.cell(row=current_row, column=3).border = border

                # Committed date of Commissioning
                sheet.cell(row=current_row, column=4).value = data.get("committed_date", "")
                sheet.cell(row=current_row, column=4).alignment = center_alignment
                sheet.cell(row=current_row, column=4).border = border

                # DC Capacity
                sheet.cell(row=current_row, column=5).value = data.get("dc_capacity", "")
                sheet.cell(row=current_row, column=5).alignment = center_alignment
                sheet.cell(row=current_row, column=5).border = border

                # AC Capacity
                sheet.cell(row=current_row, column=6).value = data.get("ac_capacity", "")
                sheet.cell(row=current_row, column=6).alignment = center_alignment
                sheet.cell(row=current_row, column=6).border = border

                # Delay of commissioning (Aging) - Yellow background
                sheet.cell(row=current_row, column=7).value = data.get("delay_days", "")
                sheet.cell(row=current_row, column=7).alignment = center_alignment
                sheet.cell(row=current_row, column=7).border = border
                sheet.cell(row=current_row, column=7).fill = yellow_fill

                # Activities Delayed
                sheet.cell(row=current_row, column=8).value = data.get("activities_delayed", "")
                sheet.cell(row=current_row, column=8).alignment = center_alignment
                sheet.cell(row=current_row, column=8).border = border

                # Revised Timelines - Revision 1
                sheet.cell(row=current_row, column=9).value = data.get("revision_1", "")
                sheet.cell(row=current_row, column=9).alignment = center_alignment
                sheet.cell(row=current_row, column=9).border = border

                # Revised Timelines - Revision 2
                sheet.cell(row=current_row, column=10).value = data.get("revision_2", "")
                sheet.cell(row=current_row, column=10).alignment = center_alignment
                sheet.cell(row=current_row, column=10).border = border

                # Revised Timelines - Revision 3
                sheet.cell(row=current_row, column=11).value = data.get("revision_3", "")
                sheet.cell(row=current_row, column=11).alignment = center_alignment
                sheet.cell(row=current_row, column=11).border = border

                # Remarks/Justification
                sheet.cell(row=current_row, column=12).value = data.get("remarks", "")
                sheet.cell(row=current_row, column=12).alignment = center_alignment
                sheet.cell(row=current_row, column=12).border = border

                sheet.row_dimensions[current_row].height = 25
                current_row += 1

            # Add empty rows (7-30)
            for row in range(current_row, 31):
                for col in range(1, 13):
                    cell = sheet.cell(row=row, column=col)
                    cell.value = ""
                    cell.alignment = center_alignment
                    cell.border = border
                    
                    # Yellow fill for column G (Aging)
                    if col == 7:
                        cell.fill = yellow_fill
                
                sheet.row_dimensions[row].height = 25

            # Freeze panes at row 6, column B
            sheet.freeze_panes = 'B6'

            # Define the file path to save the report
            reports_dir = os.path.join(settings.MEDIA_ROOT, "reports/project/")
            os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
            datetime_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(reports_dir, f"project_iar_report_{datetime_str}.xlsx")

            # Save the workbook to the file path
            workbook.save(file_path)

            # Construct the file URL
            file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/project/project_iar_report_{datetime_str}.xlsx")
            return Response({"status": True, "message": "Project IAR report generated successfully", "file_url": file_url}, status=200)


        except Exception as e:
            logger.error(f"Error generating Project Delay Analysis report: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)