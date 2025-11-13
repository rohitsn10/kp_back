import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging

logger = logging.getLogger(__name__)


def add_guvnl_sheet(workbook):
            sheet = workbook.create_sheet("GUVNL")

            # Define styles
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Blue
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            light_green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Light green
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)
            white_bold_font = Font(bold=True, size=11, color="FFFFFF")

            # Set column widths
            column_widths = {
                'B': 5, 'C': 15, 'D': 12, 'E': 15, 'F': 20, 'G': 12, 'H': 12
            }
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # ============= SECTION 1: CONNECTIVITY | HYBRID 200 MW GUVNL =============
            current_row = 2
            
            # Section header
            sheet.merge_cells(f'B{current_row}:H{current_row}')
            cell = sheet[f'B{current_row}']
            cell.value = 'CONNECTIVITY | HYBRID 200 MW GUVNL'
            cell.font = white_bold_font
            cell.fill = blue_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.row_dimensions[current_row].height = 25
            current_row += 1

            # Column headers for section 1
            headers_s1 = [
                ('B', 'S\nN'), ('C', 'Project'), ('D', 'District'), ('E', 'GSS'),
                ('F', 'Allocated Connectivity'), ('G', 'Solar'), ('H', 'Wind')
            ]
            for col_letter, text in headers_s1:
                cell = sheet[f'{col_letter}{current_row}']
                cell.value = text
                cell.font = bold_font
                cell.fill = yellow_fill if col_letter in ['G', 'H'] else blue_fill
                if col_letter in ['G', 'H']:
                    cell.font = bold_font
                else:
                    cell.font = white_bold_font
                cell.alignment = center_alignment
                cell.border = border
            sheet.row_dimensions[current_row].height = 30
            current_row += 1

            # Data row for section 1
            data_s1 = [
                ('B', '1'), ('C', 'GUVNL-Khavda'), ('D', 'Bhuj'), ('E', 'Khavda'),
                ('F', '-'), ('G', '200'), ('H', '0')
            ]
            for col_letter, value in data_s1:
                cell = sheet[f'{col_letter}{current_row}']
                cell.value = value
                cell.alignment = center_alignment
                cell.border = border
            sheet.row_dimensions[current_row].height = 25
            current_row += 2

            # ============= SECTION 2: HYBRID 50 MW GUVNL =============
            
            # Section header
            sheet.merge_cells(f'B{current_row}:H{current_row}')
            cell = sheet[f'B{current_row}']
            cell.value = 'HYBRID 50 MW GUVNL'
            cell.font = white_bold_font
            cell.fill = blue_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.row_dimensions[current_row].height = 25
            current_row += 1

            # Column headers for section 2
            headers_s2 = [
                ('B', 'S\nN'), ('C', 'Project'), ('D', 'District'), ('E', 'GSS'),
                ('F', 'Allocated Connectivity'), ('G', 'Solar'), ('H', 'Wind')
            ]
            for col_letter, text in headers_s2:
                cell = sheet[f'{col_letter}{current_row}']
                cell.value = text
                cell.font = bold_font
                cell.fill = yellow_fill if col_letter in ['G', 'H'] else blue_fill
                if col_letter in ['G', 'H']:
                    cell.font = bold_font
                else:
                    cell.font = white_bold_font
                cell.alignment = center_alignment
                cell.border = border
            sheet.row_dimensions[current_row].height = 30
            current_row += 1

            # Data row for section 2
            data_s2 = [
                ('B', '1'), ('C', 'GUVNL 50'), ('D', 'Bharuch'), ('E', '220Vagra/Vilayat'),
                ('F', '50'), ('G', '50'), ('H', '16.95')
            ]
            for col_letter, value in data_s2:
                cell = sheet[f'{col_letter}{current_row}']
                cell.value = value
                cell.alignment = center_alignment
                cell.border = border
            sheet.row_dimensions[current_row].height = 25
            current_row += 2

            # ============= SECTION 3: HYBRID 370 MW GUVNL =============
            
            # Section header with totals
            sheet.merge_cells(f'B{current_row}:E{current_row}')
            cell = sheet[f'B{current_row}']
            cell.value = 'HYBRID 370 MW GUVNL'
            cell.font = white_bold_font
            cell.fill = blue_fill
            cell.alignment = center_alignment
            cell.border = border
            
            # Totals in header
            totals_header = [('F', '386.2'), ('G', '370'), ('H', '124.2')]
            for col_letter, value in totals_header:
                cell = sheet[f'{col_letter}{current_row}']
                cell.value = value
                cell.font = bold_font
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = border
            sheet.row_dimensions[current_row].height = 25
            current_row += 1

            # Column headers for section 3
            headers_s3 = [
                ('B', 'S\nN'), ('C', 'Project'), ('D', 'District'), ('E', 'GSS'),
                ('F', 'Rev. allocated\nconnectivity'), ('G', 'Solar'), ('H', 'Wind')
            ]
            for col_letter, text in headers_s3:
                cell = sheet[f'{col_letter}{current_row}']
                cell.value = text
                cell.font = bold_font
                cell.fill = yellow_fill if col_letter in ['F', 'G', 'H'] else blue_fill
                if col_letter in ['F', 'G', 'H']:
                    cell.font = bold_font
                else:
                    cell.font = white_bold_font
                cell.alignment = center_alignment
                cell.border = border
            sheet.row_dimensions[current_row].height = 35
            current_row += 1

            # Data rows for section 3
            data_s3 = [
                {'sn': '1', 'project': 'GUVNL 370', 'district': 'Bharuch', 'gss': 'Mobha', 
                 'rev_alloc': '60', 'solar': '60', 'wind': '0'},
                {'sn': '2', 'project': 'GUVNL 370', 'district': 'Vadodara', 'gss': 'Gavasad', 
                 'rev_alloc': '70', 'solar': '70', 'wind': '0'},
                {'sn': '3', 'project': 'GUVNL 370', 'district': 'Bharuch', 'gss': 'Dahej', 
                 'rev_alloc': '70', 'solar': '70', 'wind': '45.9'},
                {'sn': '4', 'project': 'GUVNL 370', 'district': 'Bharuch', 'gss': 'Vilayat', 
                 'rev_alloc': '100', 'solar': '100', 'wind': '62.1'},
                {'sn': '5', 'project': 'GUVNL 370', 'district': 'Bharuch', 'gss': 'Vagra', 
                 'rev_alloc': '16.2', 'solar': '0', 'wind': '16.2'},
                {'sn': '6', 'project': 'GUVNL 370', 'district': 'Bharuch', 'gss': 'Achhalia', 
                 'rev_alloc': '70', 'solar': '70', 'wind': '0'},
            ]

            for data in data_s3:
                row_data = [
                    ('B', data['sn']), ('C', data['project']), ('D', data['district']), 
                    ('E', data['gss']), ('F', data['rev_alloc']), ('G', data['solar']), 
                    ('H', data['wind'])
                ]
                for col_letter, value in row_data:
                    cell = sheet[f'{col_letter}{current_row}']
                    cell.value = value
                    cell.alignment = center_alignment
                    cell.border = border
                    if col_letter in ['F', 'G', 'H']:
                        cell.fill = light_green_fill
                sheet.row_dimensions[current_row].height = 25
                current_row += 1
            
            current_row += 1

            # ============= SECTION 4: SOLAR 250 MW GUVNL =============
            
            # Section header with totals
            sheet.merge_cells(f'B{current_row}:E{current_row}')
            cell = sheet[f'B{current_row}']
            cell.value = 'SOLAR 250 MW GUVNL'
            cell.font = white_bold_font
            cell.fill = blue_fill
            cell.alignment = center_alignment
            cell.border = border
            
            # Totals
            totals_s4 = [('F', '250'), ('G', '250')]
            for col_letter, value in totals_s4:
                cell = sheet[f'{col_letter}{current_row}']
                cell.value = value
                cell.font = bold_font
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = border
            sheet.row_dimensions[current_row].height = 25
            current_row += 1

            # Column headers for section 4
            headers_s4 = [
                ('B', 'S\nN'), ('C', 'Project'), ('D', 'District'), ('E', 'GSS'),
                ('F', 'Rev. allocated\nconnectivity'), ('G', 'Solar')
            ]
            for col_letter, text in headers_s4:
                cell = sheet[f'{col_letter}{current_row}']
                cell.value = text
                cell.font = bold_font
                cell.fill = yellow_fill if col_letter in ['F', 'G'] else blue_fill
                if col_letter in ['F', 'G']:
                    cell.font = bold_font
                else:
                    cell.font = white_bold_font
                cell.alignment = center_alignment
                cell.border = border
            sheet.row_dimensions[current_row].height = 35
            current_row += 1

            # Data rows for section 4
            data_s4 = [
                {'sn': '1', 'project': 'GUVNL 250', 'district': 'Bharuch', 'gss': 'Valia', 
                 'rev_alloc': '65', 'solar': '65'},
                {'sn': '2', 'project': 'GUVNL 250', 'district': 'Bharuch', 'gss': 'Kara', 
                 'rev_alloc': '33', 'solar': '33'},
                {'sn': '3', 'project': 'GUVNL 250', 'district': 'Bharuch', 'gss': 'Randeri', 
                 'rev_alloc': '40', 'solar': '40'},
            ]

            for data in data_s4:
                row_data = [
                    ('B', data['sn']), ('C', data['project']), ('D', data['district']), 
                    ('E', data['gss']), ('F', data['rev_alloc']), ('G', data['solar'])
                ]
                for col_letter, value in row_data:
                    cell = sheet[f'{col_letter}{current_row}']
                    cell.value = value
                    cell.alignment = center_alignment
                    cell.border = border
                sheet.row_dimensions[current_row].height = 25
                current_row += 1

            # Save to response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="GUVNL_Connectivity_Report.xlsx"'
            workbook.save(response)


 