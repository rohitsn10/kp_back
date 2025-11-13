import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging

logger = logging.getLogger(__name__)


def add_66kv_project_status(workbook):
            # Create workbook and sheet
            sheet = workbook.create_sheet("66kv Project Status")

            # Define styles
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            light_blue_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # Light blue
            light_green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Light green
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            column_widths = {
                'A': 5,   # S N
                'B': 5,   # Project Type
                'C': 15,  # Project Name
                'D': 12,  # PSS Name
                'E': 12,  # GSS Name
                'F': 12,  # Project Type
                'G': 12,  # Plant DC capacity (Mwp)
                'H': 12,  # Plant AC capacity (MW)
                'I': 10,  # Wind
                'J': 12,  # Major component
                'K': 10,  # WTG Nos.
                'L': 20,  # Connectivity On the name of
                'M': 15,  # Remark
                'N': 15,  # GEDA Login Credential
                'O': 15,  # Provisional GEDA
                'P': 20,  # GETCO Connectivity Application (Stage-1)
                'Q': 25,  # Documents Required from BD/Customer for connectivity with BG
                'R': 15,  # Documents Received from Land
                'S': 20,  # GETCO Connectivity & Advocate Application (Stage-2)
                 'T': 25,  # GETCO Connectivity Approval (Stage-1)
                'U': 25,  # GETCO Connectivity Approval (Stage-2)
                'V': 20,  # Estimate Payment & Agreements
                'W': 30,  # Coordinate verification application
                'X': 30,  # Coordinate verification Report
                'Y': 35,  # Section 68 & 164 application for EHV
                'Z': 35,  # Section 68 & 164 approval for EHV
                'AA': 40, # Route Survey & other data for 68 application
                'AB': 35, # Section 68 & 164 application for 33kV OH line
                'AC': 35, # Section 68 & 164 approval for 33kV OH line
                'AD': 15, # Discom NOC
                'AE': 25, # Final GEDA Application
                'AF': 25, # Final GEDA Letter/DP
                'AG': 20, # TP application
                'AH': 20, # TP Approval
                'AI': 30, # ABT Serial Number Application
                'AJ': 30, # Meter Serial Number Letter
                'AK': 15, # Meter Supply
                'AL': 20, # ABT Meter Testing
                'AM': 25, # ABT I&C approval
                'AN': 35, # ABT Meter sealing at plan end
                'AO': 35, # ABT Meter sealing at 33kV Feeder Bay
                'AP': 30, # Metering scheme application
                'AQ': 30, # Metering scheme approval
                'AR': 25, # CEIG Plan Application
                'AS': 25, # CEIG Plan Approval
                'AT': 25, # CEIG Charging application
                'AU': 20, # CEIG Charging
                'AV': 25, # MTOA/LTOA Application
                'AW': 20, # Deficit letter from GETCO
                'AX': 30, # NOC of Generating Discom
                'AY': 20, # ABT at Factory
                'AZ': 30, # NOC of Consumer Discom
                'BA': 20, # BG & SBLC
                'BB': 25, # NOC of GETCO Finance
                'BC': 20, # NOC of SLDC
                'BD': 35, # BPTA & Other Documents submission
                'BE': 25, # BPTA approval
                'BF': 20, # RFID Details from Project
                'BG': 20, # Wheeling Signing
                'BH': 25, # GEDA COD Application
                'BI': 35, # Collection of signed Wheeling
                'BJ': 20, # F&S Start
                'BK': 35, # FTC Documents submission to SLDC
                'BL': 30, # Email reply from SLDC to GEDA for FTC
                'BM': 35, # Target Date for mail to Discom for site visit
                'BN': 40, # GEDA & Discom Site Visit & MOM
                'BO': 25, # GEDA COD Letter
                'BP': 20, # RTU with SLDC
                'BQ': 35, # WCC from GETCO Required for PSS charging
                'BR': 15, # BG Return
                'BS': 25, # Power Credit in bill
            }

            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Row 2: Column headers
            headers = [
                ('A2', 'S\nN'),
                ('B2', 'Project Type'),
                ('C2', 'Project Name'),
                ('D2', 'PSS\nName'),
                ('E2', 'GSS Name'),
                ('F2', 'Project\nType'),
                ('G2', 'Plant DC\ncapacity\n(Mwp)'),
                ('H2', 'Plant AC\ncapacity\n(MW)'),
                ('I2', 'Wind'),
                ('J2', 'Major\ncompone\nnt'),
                ('K2', 'WTG Nos.'),
                ('L2', 'Connectivity On the\nname of'),
                ('M2', 'Remark'),
                ('N2', 'GEDA\nLogin\nCredential'),
                ('O2', 'Provisiona\nl GEDA'),
                ('P2', 'GETCO Connectivity\nApplication\n(Stage-1)'),
                ('Q2', 'Documents Required from\nBD/Customer for\nconnectivity with BG'),
                ('R2', 'Documents\nReceived from Land\n& Advocate'),
                ('S2', 'GETCO Connectivity\nApplication\n(Stage-2)'),
                ('T2', 'GETCO Connectivity\nApproval\n(Stage-1)'),
                ('U2', 'GETCO Connectivity\nApproval\n(Stage-2)'),
                ('V2', 'Estimate Payment\n& Agreements'),
                ('W2', 'Coordinate\nverification\napplication'),
                ('X2', 'Coordinate\nverification\nReport'),
                ('Y2', 'Section 68 & 164\napplication\nfor EHV'),
                ('Z2', 'Section 68 & 164\napproval\nfor EHV'),
                ('AA2', 'Route Survey &\nother data for\n68 application'),
                ('AB2', 'Section 68 & 164\napplication\nfor 33kV OH line'),
                ('AC2', 'Section 68 & 164\napproval\nfor 33kV OH line'),
                ('AD2', 'Discom NOC'),
                ('AE2', 'Final GEDA\nApplication'),
                ('AF2', 'Final GEDA\nLetter/DP'),
                ('AG2', 'TP application'),
                ('AH2', 'TP Approval'),
                ('AI2', 'ABT Serial Number\nApplication'),
                ('AJ2', 'Meter Serial\nNumber Letter'),
                ('AK2', 'Meter Supply'),
                ('AL2', 'ABT Meter\nTesting'),
                ('AM2', 'ABT I&C\napproval'),
                ('AN2', 'ABT Meter\nsealing at\nplan end'),
                ('AO2', 'ABT Meter\nsealing at\n33kV Feeder Bay'),
                ('AP2', 'Metering scheme\napplication'),
                ('AQ2', 'Metering scheme\napproval'),
                ('AR2', 'CEIG Plan\nApplication'),
                ('AS2', 'CEIG Plan\nApproval'),
                ('AT2', 'CEIG Charging\napplication'),
                ('AU2', 'CEIG Charging'),
                ('AV2', 'MTOA/LTOA\nApplication'),
                ('AW2', 'Deficit letter\nfrom GETCO'),
                ('AX2', 'NOC of\nGenerating Discom'),
                ('AY2', 'ABT at\nFactory'),
                ('AZ2', 'NOC of\nConsumer Discom'),
                ('BA2', 'BG & SBLC'),
                ('BB2', 'NOC of\nGETCO Finance'),
                ('BC2', 'NOC of\nSLDC'),
                ('BD2', 'BPTA & Other\nDocuments\nsubmission'),
                ('BE2', 'BPTA approval'),
                ('BF2', 'RFID Details\nfrom Project'),
                ('BG2', 'Wheeling\nSigning'),
                ('BH2', 'GEDA COD\nApplication'),
                ('BI2', 'Collection of\nsigned Wheeling'),
                ('BJ2', 'F&S Start'),
                ('BK2', 'FTC Documents\nsubmission to SLDC'),
                ('BL2', 'Email reply from\nSLDC to GEDA\nfor FTC'),
                ('BM2', 'Target Date for\nmail to Discom\nfor site visit'),
                ('BN2', 'GEDA & Discom\nSite Visit & MOM'),
                ('BO2', 'GEDA COD\nLetter'),
                ('BP2', 'RTU with SLDC'),
                ('BQ2', 'WCC from GETCO\nRequired for\nPSS charging'),
                ('BR2', 'BG Return'),
                ('BS2', 'Power Credit\nin bill'),
            
            ]

            for cell_ref, header_text in headers:
                cell = sheet[cell_ref]
                cell.value = header_text
                cell.font = bold_font
                cell.fill = light_blue_fill
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[2].height = 60

            # Sample data
            sample_data = [
                {
                    "sn": 1,
                    "project_type": "CPP",
                    "project_name": "Aditya birla",
                    "pss_name": "66kV\nFulsar",
                    "gss_name": "220kV\nShekadar",
                    "proj_type": "Hybrid",
                    "dc_capacity": "25.60",
                    "ac_capacity": "17.20",
                    "wind": "23.10",
                    "major_component": "Wind",
                    "wtg_nos": "11.00",
                    "connectivity_name": "KPEL",
                    "remark": "Completed",
                    "geda_login": "Completed",
                    "provisional_geda": "Completed",
                    "getco_stage1": "Completed",
                    "documents_required": "Completed",
                    "documents_received": "Completed",
                    "getco_stage2": "Completed",
                }
            ]

            # Add data rows starting from row 3
            current_row = 3
            for data in sample_data:
                columns_data = [
                    (1, data["sn"]),
                    (2, data["project_type"]),
                    (3, data["project_name"]),
                    (4, data["pss_name"]),
                    (5, data["gss_name"]),
                    (6, data["proj_type"]),
                    (7, data["dc_capacity"]),
                    (8, data["ac_capacity"]),
                    (9, data["wind"]),
                    (10, data["major_component"]),
                    (11, data["wtg_nos"]),
                    (12, data["connectivity_name"]),
                    (13, data["remark"]),
                    (14, data["geda_login"]),
                    (15, data["provisional_geda"]),
                    (16, data["getco_stage1"]),
                    (17, data["documents_required"]),
                    (18, data["documents_received"]),
                    (19, data["getco_stage2"]),
                ]

                for col, value in columns_data:
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = value
                    cell.alignment = center_alignment
                    cell.border = border
                    
                    # Green fill for status columns (M to S)
                    if col >= 13 and value == "Completed":
                        cell.fill = light_green_fill

                sheet.row_dimensions[current_row].height = 30
                current_row += 1

            # Add empty rows for data entry (rows 4-10)
            for row in range(current_row, current_row + 7):
                for col in range(1, 20):
                    cell = sheet.cell(row=row, column=col)
                    cell.value = ""
                    cell.alignment = center_alignment
                    cell.border = border
                sheet.row_dimensions[row].height = 30

            # Freeze panes at row 3, column D
            sheet.freeze_panes = 'D3'

            # Save to response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="Project_Status_Tracker.xlsx"'
            workbook.save(response)

      