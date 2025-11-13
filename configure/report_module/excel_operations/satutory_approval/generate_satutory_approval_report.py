import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging
from .guvnl_sheet import add_guvnl_sheet
from .project_status_66kv import add_66kv_project_status
from datetime import datetime
logger = logging.getLogger(__name__)


def generate_66kv_satutory_status_report(request):
        try:
            # Create workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "66kV Connectivity"

            # Define styles
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            column_widths = {
                'A': 8,   # Sr No
                'B': 15,  # Project area
                'C': 30,  # Name of Company
                'D': 15,  # GETCO S/S
                'E': 12,  # Total approved Evacuation
                'F': 12,  # Under approval
                'G': 18,  # Evacuation/Connectivity Type
                'H': 15,  # Allocations (in MW)
                'I': 50,  # Solar
                'J': 10,  # Wind
                'K': 15,  # Balance allocation
                'L': 20,  # PSS Charging Date & status
                'M': 20,  # 10% Charging Date (Tentative)
                'N': 15,  # 100% charging
            }

            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Row 1: Main title
            sheet.merge_cells('A1:N1')
            cell = sheet['A1']
            cell.value = 'RE Current Connectivity- 66kV'
            cell.font = Font(bold=True, size=14)
            cell.fill = orange_fill
            cell.alignment = center_alignment
            cell.border = border
            sheet.row_dimensions[1].height = 30

            # Row 2: Column headers
            headers = [
                ('A2', 'Sr\nNo'),
                ('B2', 'Project area'),
                ('C2', 'Name of Company'),
                ('D2', 'GETCO S/S'),
                ('E2', 'Total\napproved\nEvacuation'),
                ('F2', 'Under\napproval'),
                ('G2', 'Evacuation/C\nonnectivity\nType'),
                ('H2', 'Allocations\n(in MW)'),
                ('I2', 'Solar'),
                ('J2', 'Wind'),
                ('K2', 'Balance\nallocatio\nn'),
                ('L2', 'PSS Charging\nDate &\nstatus'),
                ('M2', '10% Charging Date\n(Tentative)'),
                ('N2', '100\ncha'),
            ]

            for cell_ref, header_text in headers:
                cell = sheet[cell_ref]
                cell.value = header_text
                cell.font = bold_font
                cell.fill = yellow_fill
                cell.alignment = center_alignment
                cell.border = border

            sheet.row_dimensions[2].height = 50

            # Sample data
            sample_data = [
                {
                    "sr": 1,
                    "area": "Bharuch",
                    "companies": [
                        {"name": "KPI Green Energy Limited", "getco": "66KV Hansot", "total": "30", "under": "", 
                         "evac_type": "Solar", "alloc": "3.9", "solar": "For CPP Use, allocated 3.9 MW to Bhathwari", 
                         "wind": "NA", "balance": "26.1", "pss": "June'25", "charging_10": "", "charging_100": ""},
                        {"name": "Sun Drops Energia Pvt Ltd", "getco": "66KV Vagra", "total": "", "under": "", 
                         "evac_type": "", "alloc": "6.05", "solar": "6.05", 
                         "wind": "5.6", "balance": "", "pss": "", "charging_10": "", "charging_100": ""},
                    ]
                },
                {
                    "sr": 2,
                    "area": "Bharuch",
                    "companies": [
                        {"name": "Anupam Rasayan India Ltd.", "getco": "66KV Vagra", "total": "", "under": "", 
                         "evac_type": "", "alloc": "5.6", "solar": "3.025", 
                         "wind": "5.6", "balance": "", "pss": "", "charging_10": "", "charging_100": ""},
                        {"name": "CTX Life Science", "getco": "66KV Vagra", "total": "60", "under": "", 
                         "evac_type": "Hybrid", "alloc": "2.1", "solar": "0.825", 
                         "wind": "2.1", "balance": "1.45", "pss": "PSS Charged", "charging_10": "Done", "charging_100": ""},
                        {"name": "KPEL", "getco": "66KV Vagra", "total": "", "under": "", 
                         "evac_type": "", "alloc": "28.6", "solar": "0", 
                         "wind": "28.6", "balance": "", "pss": "", "charging_10": "", "charging_100": ""},
                    ]
                },
                {
                    "sr": 3,
                    "area": "Bharuch",
                    "companies": [
                        {"name": "KPI Green Energy Limited", "getco": "66KV Vagra", "total": "", "under": "", 
                         "evac_type": "", "alloc": "16.2", "solar": "For GUVNL-2 (Wind Part)", 
                         "wind": "16.2", "balance": "", "pss": "", "charging_10": "", "charging_100": ""},
                        {"name": "Sun Drops Energia Pvt Ltd", "getco": "66KV Vagra", "total": "", "under": "", 
                         "evac_type": "", "alloc": "0.44", "solar": "9.65", 
                         "wind": "NA", "balance": "", "pss": "", "charging_10": "", "charging_100": ""},
                        {"name": "KPI Green Energy Limited", "getco": "66KV Vagra", "total": "30", "under": "", 
                         "evac_type": "Solar", "alloc": "4.4", "solar": "4.4", 
                         "wind": "NA", "balance": "9.21", "pss": "PSS Charged", "charging_10": "Done", "charging_100": ""},
                        {"name": "Creative Technology", "getco": "66KV Vagra", "total": "", "under": "", 
                         "evac_type": "", "alloc": "9.9", "solar": "9.9", 
                         "wind": "NA", "balance": "", "pss": "", "charging_10": "", "charging_100": ""},
                        {"name": "Gujarat Polyfilms", "getco": "66KV Vagra", "total": "", "under": "", 
                         "evac_type": "", "alloc": "6.05", "solar": "6.05", 
                         "wind": "NA", "balance": "", "pss": "", "charging_10": "", "charging_100": ""},
                    ]
                }
            ]

            # Add data rows starting from row 3
            current_row = 3
            for data in sample_data:
                start_row = current_row
                
                # Add all company rows
                for idx, company in enumerate(data["companies"]):
                    # Sr No (only for first row of each group)
                    if idx == 0:
                        cell = sheet.cell(row=current_row, column=1)
                        cell.value = data["sr"]
                        cell.alignment = center_alignment
                        cell.border = border
                    else:
                        cell = sheet.cell(row=current_row, column=1)
                        cell.value = ""
                        cell.border = border

                    # Project area (only for first row of each group)
                    if idx == 0:
                        cell = sheet.cell(row=current_row, column=2)
                        cell.value = data["area"]
                        cell.alignment = center_alignment
                        cell.border = border
                    else:
                        cell = sheet.cell(row=current_row, column=2)
                        cell.value = ""
                        cell.border = border

                    # Company details
                    columns_data = [
                        (3, company["name"], left_alignment),
                        (4, company["getco"], center_alignment),
                        (5, company["total"], center_alignment),
                        (6, company["under"], center_alignment),
                        (7, company["evac_type"], center_alignment),
                        (8, company["alloc"], center_alignment),
                        (9, company["solar"], left_alignment),
                        (10, company["wind"], center_alignment),
                        (11, company["balance"], center_alignment),
                        (12, company["pss"], center_alignment),
                        (13, company["charging_10"], center_alignment),
                        (14, company["charging_100"], center_alignment),
                    ]

                    for col, value, alignment in columns_data:
                        cell = sheet.cell(row=current_row, column=col)
                        cell.value = value
                        cell.alignment = alignment
                        cell.border = border

                    sheet.row_dimensions[current_row].height = 25
                    current_row += 1

                # Merge Sr No and Project area cells for the group
                if len(data["companies"]) > 1:
                    sheet.merge_cells(f'A{start_row}:A{current_row-1}')
                    sheet.merge_cells(f'B{start_row}:B{current_row-1}')

            # Freeze panes at row 3, column C
            sheet.freeze_panes = 'C3'
            add_guvnl_sheet(workbook)  # Add GUVNL sheet
            add_66kv_project_status(workbook)
            # Prepare HTTP response
                 # Define the file path to save the report
            reports_dir = os.path.join(settings.MEDIA_ROOT, "reports/satutory_approval/")
            os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
            datetime_str= datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(reports_dir, f"66kv_satutory_status_report_{datetime_str}.xlsx")

            # Save the workbook to the file path
            workbook.save(file_path)


            file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/satutory_approval/66kv_satutory_status_report_{datetime_str}.xlsx")

            return Response({"file_url": file_url,"status":True,"message":"Satutory Approval 66KV report generated successfully"},status=200)

        except Exception as e:
            logger.error(f"Error generating RE Current Connectivity 66kV report: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)