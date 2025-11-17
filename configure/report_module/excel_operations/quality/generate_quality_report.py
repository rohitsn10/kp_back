import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging
import os
from django.conf import settings
from datetime import datetime
from .safty_doc import generate_safty_doc_report
logger = logging.getLogger(__name__)


def generate_quality_report(request):
        try:
            # Create workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Quality Doc"

            # Define styles
            header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # Light blue
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            sheet.column_dimensions['A'].width = 8   # SI No
            sheet.column_dimensions['B'].width = 45  # Description of Check List
            sheet.column_dimensions['C'].width = 25  # Type of Check List/Register/Reports
            sheet.column_dimensions['D'].width = 15  # Department
            sheet.column_dimensions['E'].width = 12  # Revision
            sheet.column_dimensions['F'].width = 15  # Remarks

            # Row 2: Headers
            headers = [
                ('A2', 'SI No'),
                ('B2', 'Description of Check List'),
                ('C2', 'Type of Check\nList/Register/Reports'),
                ('D2', 'Department'),
                ('E2', 'Revision'),
                ('F2', 'Remarks'),
            ]

            for cell_ref, header_text in headers:
                cell = sheet[cell_ref]
                cell.value = header_text
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Set row height for header
            sheet.row_dimensions[2].height = 35

            # Checklist data - you can fetch this from database
            checklist_data = [
                {"sl": 1, "description": "Asphalt Road", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 2, "description": "Boundary Fencing", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 3, "description": "Brick Masonry", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 4, "description": "Brick Plastering", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 5, "description": "Building Painting", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 6, "description": "Cement Register", "type": "Register", "department": "Civil", "revision": "R-00"},
                {"sl": 7, "description": "Compressive Strength of Bricks", "type": "Report", "department": "Civil", "revision": "R-00"},
                {"sl": 8, "description": "Compressive Strength of Concrete Cube", "type": "Report", "department": "Civil", "revision": "R-00"},
                {"sl": 9, "description": "Concrete Pour Card", "type": "Check List/Register", "department": "Civil", "revision": "R-00"},
                {"sl": 10, "description": "Cube Testing Report Master Card", "type": "Register", "department": "Civil", "revision": "R-00"},
                {"sl": 11, "description": "Curing Log Sheet", "type": "Check List/Register", "department": "Civil", "revision": "R-00"},
                {"sl": 12, "description": "Determination of Silt Content in Fine Aggregate", "type": "Report", "department": "Civil", "revision": "R-00"},
                {"sl": 13, "description": "Drainage", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 14, "description": "Excavation", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 15, "description": "Fencing Pole Concrete", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 16, "description": "ICR Structure Installation", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 17, "description": "Land Levelling", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 18, "description": "Main Gate", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 19, "description": "Moisture Content of Coarse & Fine Aggregate", "type": "Report", "department": "Civil", "revision": "R-00"},
                {"sl": 20, "description": "PCC", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 21, "description": "Pile Cap Concrete", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 22, "description": "Pile Concreting Inspection", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 23, "description": "Pour Card for Column Concrete", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 24, "description": "Pour Card for Footing Concrete", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 25, "description": "Pour Card for Plinth Beam-Lintel Beam-Roof Beam", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 26, "description": "Pour Card for Slab Concrete", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 27, "description": "Sieve Analysis of Coarse Aggregate Report", "type": "Report", "department": "Civil", "revision": "R-00"},
                {"sl": 28, "description": "Sieve Analysis of Fine Aggregate Report", "type": "Report", "department": "Civil", "revision": "R-00"},
                {"sl": 29, "description": "Transformer Fencing Installation", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 30, "description": "Water Absorption of Bricks", "type": "Report", "department": "Civil", "revision": "R-00"},
                {"sl": 31, "description": "WBM Road", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 32, "description": "MCS Pumping System", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 33, "description": "MCS RO Plant Installation", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 34, "description": "MCS Pipe Laying", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 35, "description": "Module Mounting", "type": "Check List", "department": "Mechanical", "revision": "R-00"},
                {"sl": 36, "description": "Module Mounting Structure", "type": "Check List", "department": "Mechanical", "revision": "R-00"},
                {"sl": 37, "description": "AC Cable Laying", "type": "Check List", "department": "Electrical", "revision": "R-00"},
                {"sl": 38, "description": "Cable trench (AC andd DC)", "type": "Check List", "department": "Electrical", "revision": "R-00"},
            ]

            # Add data rows starting from row 3
            current_row = 3
            for item in checklist_data:
                # SI No
                cell = sheet.cell(row=current_row, column=1)
                cell.value = item["sl"]
                cell.alignment = center_alignment
                cell.border = border

                # Description
                cell = sheet.cell(row=current_row, column=2)
                cell.value = item["description"]
                cell.alignment = left_alignment
                cell.border = border

                # Type of Check List/Register/Reports
                cell = sheet.cell(row=current_row, column=3)
                cell.value = item["type"]
                cell.alignment = center_alignment
                cell.border = border

                # Department
                cell = sheet.cell(row=current_row, column=4)
                cell.value = item["department"]
                cell.alignment = center_alignment
                cell.border = border

                # Revision
                cell = sheet.cell(row=current_row, column=5)
                cell.value = item["revision"]
                cell.alignment = center_alignment
                cell.border = border

                # Remarks (empty)
                cell = sheet.cell(row=current_row, column=6)
                cell.value = ""
                cell.alignment = center_alignment
                cell.border = border

                # Set row height
                sheet.row_dimensions[current_row].height = 20

                current_row += 1
    
        
            generate_safty_doc_report(workbook)
            # Define the file path to save the report
            reports_dir = os.path.join(settings.MEDIA_ROOT, "reports")
            os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
            datetime_str= datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(reports_dir, f"quality_report_{datetime_str}.xlsx")

            # Save the workbook to the file path
            workbook.save(file_path)


            file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/quality_report_{datetime_str}.xlsx")

            return Response({"file_url": file_url},status=200)

        except Exception as e:
            logger.error(f"Error generating Checklist report: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)