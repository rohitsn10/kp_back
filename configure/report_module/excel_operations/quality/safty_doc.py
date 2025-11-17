import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging
import os
from django.conf import settings
from datetime import datetime

logger = logging.getLogger(__name__)


def generate_safty_doc_report(workbook):
        try:
            # Create workbook and sheet
            sheet = workbook.create_sheet("Quality Doc")

            # Define styles
            header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # Light blue
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            sheet.column_dimensions['A'].width = 8   # SI No
            sheet.column_dimensions['B'].width = 45  # Description of Check List
            sheet.column_dimensions['C'].width = 15  # Remarks

            # Row 2: Headers
            headers = [
                ('A2', 'SI No'),
                ('B2', 'Description of Check List'),
                ('C2', 'Remarks'),
            ]

            for cell_ref, header_text in headers:
                cell = sheet[cell_ref]
                cell.value = header_text
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Set row height for header
            sheet.row_dimensions[2].height = 35

            # Checklist data - you can fetch this from database
            checklist_data = [
                {"sl": 1, "description": "Asphalt Road", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 2, "description": "Boundary Fencing", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 3, "description": "Brick Masonry", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 4, "description": "Brick Plastering", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 5, "description": "Building Painting", "type": "Check List", "department": "Civil", "revision": "R-00"},
                {"sl": 6, "description": "Cement Register", "type": "Register", "department": "Civil", "revision": "R-00"},
                {"sl": 7, "description": "Compressive Strength of Bricks", "type": "Report", "department": "Civil", "revision": "R-00"},
                {"sl": 8, "description": "Compressive Strength of Concrete Cube", "type": "Report", "department": "Civil", "revision": "R-00"},
                {"sl": 9, "description": "Concrete Pour Card", "type": "Check List/Register", "department": "Civil", "revision": "R-00"},
                {"sl": 10, "description": "Cube Testing Report Master Card", "type": "Register", "department": "Civil", "revision": "R-00"},
                {"sl": 11, "description": "Curing Log Sheet", "type": "Check List/Register", "department": "Civil", "revision": "R-00"}
            ]

            # Add data rows starting from row 3
            current_row = 3
            for item in checklist_data:
                # SI No
                cell = sheet.cell(row=current_row, column=1)
                cell.value = item["sl"]
                cell.alignment = center_alignment
                cell.border = border

                # Description
                cell = sheet.cell(row=current_row, column=2)
                cell.value = item["description"]
                cell.alignment = left_alignment
                cell.border = border


                # Remarks (empty)
                cell = sheet.cell(row=current_row, column=3)
                cell.value = ""
                cell.alignment = center_alignment
                cell.border = border

                # Set row height
                sheet.row_dimensions[current_row].height = 20

                current_row += 1

        
            return Response({'workbook': workbook}, status=200)
        except Exception as e:
            logger.error(f"Error generating Checklist report: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)