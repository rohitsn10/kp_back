from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse

from django.conf import settings
from rest_framework.response import Response
import logging

logger = logging.getLogger(__name__)

def generate_scm_material_tracking_report(request):
        try:
            # Create workbook and sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Procurement Tracker"

            # Define styles
            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
            category_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light orange
            delivery_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # Light blue
            
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)

            # Set column widths
            column_widths = {
                'A': 8,   # Sr No.
                'B': 15,  # Categories
                'C': 20,  # Name of Project
                'D': 20,  # Major Items
                'E': 15,  # Project Material
                'F': 25,  # Specification of Materials
                'G': 12,  # EPC
                'H': 12,  # Quantity
                'I': 10,  # UOM
                'J': 15,  # PO No
                'K': 15,  # KP to EPC
                'L': 15,  # Payment Status
                'M': 20,  # Supplier with Contact
                'N': 12,  # March-25
                'O': 12,  # April-25
                'P': 12,  # May-25
                'Q': 12,  # June-25
                'R': 15,  # Inspection Done
                'S': 16,  # MDCC from KP
                'T': 17,  # Delivered
                'U': 18,  # Pending Material
            }

            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Row 3: Merge cells for "Delivery dates as per Vendor"
            sheet.merge_cells('N3:Q3')
            cell_delivery = sheet['N3']
            cell_delivery.value = 'Delivery dates as per Vendor'
            cell_delivery.font = bold_font
            cell_delivery.fill = delivery_fill
            cell_delivery.alignment = center_alignment
            cell_delivery.border = border

            sheet.merge_cells('V4:W4')
            cell_delivery = sheet['V4']
            cell_delivery.value = 'Remarks'
            cell_delivery.font = bold_font
            cell_delivery.fill = header_fill
            cell_delivery.alignment = center_alignment
            cell_delivery.border = border

            # Row 4: Main headers
            headers = [
                ('A4', 'Sr No.'),
                ('B4', 'Categories'),
                ('C4', 'Name of Project'),
                ('D4', 'Major Items'),
                ('E4', 'Project Material'),
                ('F4', 'Specification of Materials'),
                ('G4', 'EPC'),
                ('H4', 'Quantity'),
                ('I4', 'UOM'),
                ('J4', 'PO No'),
                ('K4', 'KP to EPC'),
                ('L4', 'Payment Status'),
                ('M4', 'Supplier with Contact'),
                ('N4', 'March-25'),
                ('O4', 'April-25'),
                ('P4', 'May-25'),
                ('Q4', 'June-25'),
                ('R4', 'Inspection Done'),
                ('S4', 'MDCC from KP'),
                ('T4', 'Delivered'),
                ('U4', 'Pending Material'),

            ]

            for cell_ref, header_text in headers:
                cell = sheet[cell_ref]
                cell.value = header_text
                cell.font = bold_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Predefined categories data
            categories = [
                'Utility - IPF',
                'Utility - CP',
                'C&I - Kusu',
                'C&I - DRE',
                'Utility - IPF',
                'C&I - CPP',
                'C&I - Kusu',
            ]

            # Add category rows with dropdowns
            current_row = 5
            for idx, category in enumerate(categories, start=1):
                # Sr No.
                sheet.cell(row=current_row, column=1).value = idx
                sheet.cell(row=current_row, column=1).alignment = center_alignment
                sheet.cell(row=current_row, column=1).border = border

                # Category with light orange fill
                cat_cell = sheet.cell(row=current_row, column=2)
                cat_cell.value = category
                cat_cell.fill = category_fill
                cat_cell.alignment = Alignment(horizontal='left', vertical='center')
                cat_cell.border = border

                # Empty cells for other columns
                for col in range(3, 19):  # Columns C to R
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = ''
                    cell.alignment = center_alignment
                    cell.border = border

                current_row += 1

            # Add some empty rows (7-16 as shown in image)
            for row in range(current_row, current_row + 10):
                for col in range(1, 19):
                    cell = sheet.cell(row=row, column=col)
                    cell.value = ''
                    cell.alignment = center_alignment
                    cell.border = border

            # Set row heights
            sheet.row_dimensions[3].height = 25
            sheet.row_dimensions[4].height = 30
            for row in range(5, current_row + 10):
                sheet.row_dimensions[row].height = 20

            # Freeze panes (freeze first 4 rows and first 2 columns)
            sheet.freeze_panes = 'C5'

          
            # Define the file path to save the report
            reports_dir = os.path.join(settings.MEDIA_ROOT, "reports")
            os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
            datetime_str= datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(reports_dir, f"scm_material_tracking_report_{datetime_str}.xlsx")

            # Save the workbook to the file path
            workbook.save(file_path)


            file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/scm_material_tracking_report_{datetime_str}.xlsx")


            return file_url

        except Exception as e:
            logger.error(f"Error generating Procurement Tracker report: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)