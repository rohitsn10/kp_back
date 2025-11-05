import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from rest_framework.response import Response
import os
from django.conf import settings
from datetime import datetime
import logging
from .data_insert import data_insert_into_row
logger = logging.getLogger(__name__)


def generate_hse_mis_report(request):
    try:
        # Create workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "HSE MIS Report"

        # Define styles
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold_font = Font(bold=True)

        # Set column widths
        sheet.column_dimensions['A'].width = 10  # Sr. No.
        sheet.column_dimensions['B'].width = 50  # Leading Parameters
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            sheet.column_dimensions[col].width = 10  # Month columns
       
        # Row 3: Sr. No. and Site Name
        sheet.merge_cells('A3:A3')
        sheet['A3'] = 'Sr. No.'
        sheet['A3'].font = bold_font
        sheet['A3'].alignment = center_alignment
        sheet['A3'].border = border

        sheet.merge_cells('B3:B3')
        sheet['B3'] = 'Site Name:-'
        sheet['B3'].font = bold_font
        sheet['B3'].alignment = Alignment(horizontal='left', vertical='center')
        sheet['B3'].border = border

        # Row 4: Leading Parameters header with months
        sheet.merge_cells('A4:B4')
        sheet['A4'] = 'Leading Parameters'
        sheet['A4'].font = bold_font
        sheet['A4'].fill = header_fill
        sheet['A4'].alignment = center_alignment
        sheet['A4'].border = border

        sheet['O4'] = 'FY-25'
        sheet['O4'].font = bold_font
        sheet['O4'].fill = header_fill
        sheet['O4'].alignment = center_alignment
        sheet['O4'].border = border

        sheet['P4'] = 'Remarks'
        sheet['P4'].font = bold_font
        sheet['P4'].fill = header_fill
        sheet['P4'].alignment = center_alignment
        sheet['P4'].border = border

        # Month headers
        months = ['Apr-24', 'May-24', 'Jun-24', 'Jul-24', 'Aug-24', 'Sep-24', 
                  'Oct-24', 'Nov-24', 'Dec-24', 'Jan-25', 'Feb-25', 'Mar-25']
        
        for idx, month in enumerate(months, start=3):  # Start from column C (index 3)
            cell = sheet.cell(row=4, column=idx)
            cell.value = month
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # Define the parameters with their data
        parameters = [
            {"sr": 1, "name": "No. of Manday's Worked", "data": ''},
            {"sr": 2, "name": "No. of Employee Worked (On Roll)", "data": ''},
            {"sr": 3, "name": "No. of Employee Worked ( Off Roll+ Contractual Manpower)", "data": ''},
            {"sr": 4, "name": "No. of Total Employee Worked", "data": [0] * 12},
            {"sr": 5, "name": "Manhours Worked Employee(On Roll)", "data": ''},
            {"sr": 6, "name": "Manhours Employee Worked (Off Roll+ Contractor Workers)", "data": ''},
            {"sr": 7, "name": "No. of Total Manhours", "data": ''},
            {"sr": 8, "name": "No. of Training Session Conducted", "data": ''},
            {"sr": 9, "name": "No. of Total Safety Training Man Hours", "data": ''},
            {"sr": 10, "name": "No. of Safety training manhours (Employee)", "data": ''},
            {"sr": 11, "name": "No. of Safety training manhours (Off Roll+ Contractor Workers)", "data": ''},
            {"sr": 12, "name": "No's of Unsafe Act Reported", "data": ''},
            {"sr": 13, "name": "No's of Unsafe Act Closed", "data": ''},
            {"sr": 14, "name": "Nos. of Unsafe Condition Reported", "data": ''},
            {"sr": 15, "name": "Nos. of Unsafe Condition Closed", "data": ''},
            {"sr": 16, "name": "No's of Near Miss Reported", "data": ''},
            {"sr": 17, "name": "No's of HSE induction", "data": ''},
            {"sr": 18, "name": "No's of PTW Issued", "data": ''},
            {"sr": 19, "name": "No.of TBT Conducted", "data": ''},
            {"sr": 20, "name": "Nos. of Safety Audits", "data": ''},
            {"sr": 21, "name": "No's of Mock Drill", "data": ''},
            {"sr": 22, "name": "No's of Reward & Recognition (R&R)", "data": ''},
            {"sr": 23, "name": "Safety Committee Meetings", "data": ''},
            {"sr": 24, "name": "Equipment & tool Inspection", "data": ''},
            {"sr": 25, "name": "Nos. of Vehicle Inspection", "data": ''},
        ]

        # Add parameter rows starting from row 5
        current_row = 5
        current_row = data_insert_into_row(current_row, parameters, sheet, border, center_alignment, Alignment)
        
        # Add Lagging Parameters header
        sheet.merge_cells(f'A{current_row}:B{current_row}')
        sheet[f'A{current_row}'] = 'Lagging Parameters'
        sheet[f'A{current_row}'].font = bold_font
        sheet[f'A{current_row}'].fill = header_fill
        sheet[f'A{current_row}'].alignment = center_alignment
        sheet[f'A{current_row}'].border = border
        
        current_row += 1

        lagging_parameters = [
            {"sr": 1, "name": "Occupational Illness", "data": ''},
            {"sr": 2, "name": "First Aid Cases (FAC)", "data": ''},
            {"sr": 3, "name": "Medical Treatment Case (MTC)", "data": ''},
            {"sr": 4, "name": "Restricted Work Cases (RWC)", "data": ''},
            {"sr": 5, "name": "Nos. of Fire Incident or Explosion", "data": ''},
            {"sr": 6, "name": "No of Lost Time Injury (LTI)", "data": ''},
            {"sr": 7, "name": "Nos. of Fatalities", "data": ''},
            {"sr": 8, "name": "Lost Time Injury Frequency Rate (LTIFR)", "data": ''},
            {"sr": 9, "name": "Total Recordable Injury Rate (TRIR)", "data": ''},
            {"sr": 10, "name": "No. of (Fines Paid) Consequences Managements", "data": ''},
        ]

        current_row = data_insert_into_row(current_row, lagging_parameters, sheet, border, center_alignment, Alignment)

        sheet.merge_cells(f'A{current_row}:B{current_row}')
        sheet[f'A{current_row}'] = 'ENV Parameter'
        sheet[f'A{current_row}'].font = bold_font
        sheet[f'A{current_row}'].fill = header_fill
        sheet[f'A{current_row}'].alignment = center_alignment
        sheet[f'A{current_row}'].border = border
        
        current_row += 1
        env_parameters = [
            {"sr": 1, "name": "Water Consumption (In Ltr)", "data": ''},
            {"sr": 2, "name": "Electricity Consumption (KWH)", "data": ''},
            {"sr": 3, "name": "Electricity Generation / Sold (KWH)", "data": ''},
        ]
        current_row = data_insert_into_row(current_row, env_parameters, sheet, border, center_alignment, Alignment)


        sheet.merge_cells(f'A{current_row}:B{current_row}')
        sheet[f'A{current_row}'] = 'Non-Hazardous Waste'
        sheet[f'A{current_row}'].font = bold_font
        sheet[f'A{current_row}'].fill = header_fill
        sheet[f'A{current_row}'].alignment = center_alignment
        sheet[f'A{current_row}'].border = border
        
        current_row += 1
        non_hazardous_waste_parameters = [
            {"sr": 1, "name": "Wood (Kg)", "data": ''},
            {"sr": 2, "name": "Metal waste (Kg)", "data": ''},
            {"sr": 3, "name": "Cartoon waste (Kg)", "data": ''},
            {"sr": 4, "name": "Plastic scrap (Kg)", "data": ''},
            {"sr": 5, "name": "Other scrap (Kg)", "data": ''},
        ]
        current_row = data_insert_into_row(current_row, non_hazardous_waste_parameters, sheet, border, center_alignment, Alignment)
        

        sheet.merge_cells(f'A{current_row}:B{current_row}')
        sheet[f'A{current_row}'] = 'Hazardous Waste'
        sheet[f'A{current_row}'].font = bold_font
        sheet[f'A{current_row}'].fill = header_fill
        sheet[f'A{current_row}'].alignment = center_alignment
        sheet[f'A{current_row}'].border = border
        
        current_row += 1
        hazardous_waste_parameters = [
            {"sr": 1, "name": "Used oil (Kg)", "data": ''},
            {"sr": 2, "name": "Oil Soak cotton waste (KG= nos *den) (Kg)", "data": ''},
            {"sr": 3, "name": "Empty Oil drum / Barrels (Kg)", "data": ''},
        ]
        current_row = data_insert_into_row(current_row, hazardous_waste_parameters, sheet, border, center_alignment, Alignment)
        
        sheet.merge_cells(f'A{current_row}:B{current_row}')
        sheet[f'A{current_row}'] = 'Fuel Consumption'
        sheet[f'A{current_row}'].font = bold_font
        sheet[f'A{current_row}'].fill = header_fill
        sheet[f'A{current_row}'].alignment = center_alignment
        sheet[f'A{current_row}'].border = border
        
        current_row += 1
        fuel_consumption_parameters = [
            {"sr": 1, "name": "Diesel Consumption (L)", "data": ''},
            {"sr": 2, "name": "Petrol Consumption (L)", "data": ''},
        ]
        current_row = data_insert_into_row(current_row, fuel_consumption_parameters, sheet, border, center_alignment, Alignment)
        
        
        sheet.merge_cells(f'A{current_row}:B{current_row}')
        sheet[f'A{current_row}'] = 'E-Waste'
        sheet[f'A{current_row}'].font = bold_font
        sheet[f'A{current_row}'].fill = header_fill
        sheet[f'A{current_row}'].alignment = center_alignment
        sheet[f'A{current_row}'].border = border
        
        current_row += 1
        e_waste_parameters = [
            {"sr": 1, "name": "Modules (Kg)", "data": ''},
            {"sr": 2, "name": "Tube light/Halogen (Kg)", "data": ''},
            {"sr": 3, "name": "Cables/MC4 Connector (Kg)", "data": ''},
            {"sr": 4, "name": "Battery (Kg)", "data": ''},
            {"sr": 5, "name": "PCB/MCB/Switches (Kg)", "data": ''},
        ]
        current_row = data_insert_into_row(current_row, e_waste_parameters, sheet, border, center_alignment, Alignment)
        # Define the file path to save the report
        reports_dir = os.path.join(settings.MEDIA_ROOT, "reports")
        os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
        datetime_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(reports_dir, f"hse_mis_report_{datetime_str}.xlsx")

        workbook.save(file_path)

        file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/hse_mis_report_{datetime_str}.xlsx")

        return file_url

    except Exception as e:
        logger.error(f"Error generating HSE MIS report: {e}", exc_info=True)
        return None