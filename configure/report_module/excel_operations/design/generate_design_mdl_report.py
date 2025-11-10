import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from rest_framework.response import Response
import logging

from django.conf import settings
import os
from datetime import datetime

logger = logging.getLogger(__name__)


def generate_design_mdl_report(request):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Design MDL Report"

            # Define styles
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Color fills
            blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Blue
            orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            light_green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Light green
            green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            pink_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light pink
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            bold_font = Font(bold=True, size=11)
            white_bold_font = Font(bold=True, size=11, color="FFFFFF")

            # Set column widths
            column_widths = {
                'A': 8, 'B': 12, 'C': 10, 'D': 10, 'E': 25, 'F': 40,
                'G': 15, 'H': 20, 'I': 12, 'J': 40, 'K': 18, 'L': 18,
                'M': 18, 'N': 18
            }
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # Row 1: Project Title
            sheet.merge_cells('A1:I1')
            cell = sheet['A1']
            cell.value = 'PROJECT: 300 MW GROUND MOUNTED SOLAR PV PLANT AT KHAVDA, GUJARAT'
            cell.font = bold_font
            cell.alignment = left_alignment
            cell.border = border
            sheet.row_dimensions[1].height = 25

            # Row 2: Developer and Document Review Summary
            sheet.merge_cells('A2:E2')
            cell = sheet['A2']
            cell.value = 'DEVELOPER:- COAL INDIA LIMITED (CIL)'
            cell.font = bold_font
            cell.alignment = left_alignment
            cell.border = border

            sheet.merge_cells('F2:I2')
            cell = sheet['F2']
            cell.value = 'Documents Review Summary of Solar PV Project'
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = border
            sheet.row_dimensions[2].height = 25

            # Row 3: EPC and Summary headers
            sheet.merge_cells('A3:E3')
            cell = sheet['A3']
            cell.value = 'EPC:- KPI GREEN ENERGY LIMITED'
            cell.font = bold_font
            cell.alignment = left_alignment
            cell.border = border

            headers_row3 = [
                ('F3', 'Description', orange_fill),
                ('G3', 'Legends', yellow_fill),
                ('H3', 'No. of Drawings', yellow_fill),
                ('I3', '% Progress', yellow_fill)
            ]
            for cell_ref, text, fill in headers_row3:
                cell = sheet[cell_ref]
                cell.value = text
                cell.font = bold_font
                cell.fill = fill
                cell.alignment = center_alignment
                cell.border = border
            sheet.row_dimensions[3].height = 25

            # Rows 4-7: Document summary
            summary_data = [
                (4, 'Total Documents', '155', 'Approved', 'A', '52', '33.5%', light_green_fill),
                (5, 'Document Submitted', '81', 'Submitted_R0', 'S', '24', '15.5%', yellow_fill),
                (6, 'Balance (Not submitted)', '74', 'Commented', 'C', '5', '3.2%', red_fill),
                (7, '', '', 'Not Submitted', 'N', '74', '47.7%', None),
            ]

            for row_num, label, count, desc, legend, num_drawings, progress, fill_color in summary_data:
                # Left side (A-C for label, D-E for count)
                if label:  # If there's text
                    # Set value first, then merge
                    cell = sheet[f'A{row_num}']
                    cell.value = label
                    cell.font = bold_font
                    cell.alignment = left_alignment
                    cell.border = border
                    sheet.merge_cells(f'A{row_num}:C{row_num}')
                    
                    # Count value
                    cell = sheet[f'D{row_num}']
                    cell.value = count
                    cell.alignment = center_alignment
                    cell.border = border
                    sheet.merge_cells(f'D{row_num}:E{row_num}')

                # Right side summary (F-I)
                cell = sheet[f'F{row_num}']
                cell.value = desc
                cell.alignment = center_alignment
                cell.border = border
                if fill_color:
                    cell.fill = fill_color
                
                cell = sheet[f'G{row_num}']
                cell.value = legend
                cell.alignment = center_alignment
                cell.border = border
                if fill_color:
                    cell.fill = fill_color
                
                cell = sheet[f'H{row_num}']
                cell.value = num_drawings
                cell.alignment = center_alignment
                cell.border = border
                if fill_color:
                    cell.fill = fill_color
                
                cell = sheet[f'I{row_num}']
                cell.value = progress
                cell.alignment = center_alignment
                cell.border = border
                if fill_color:
                    cell.fill = fill_color
                
                sheet.row_dimensions[row_num].height = 25

            # Row 9: Column number headers with Revision labels
            sheet.merge_cells('A9:A9')
            sheet['A9'].value = '1'
            sheet.merge_cells('B9:C9')
            sheet['B9'].value = '2'
            sheet.merge_cells('D9:E9')
            sheet['D9'].value = '3'
            sheet['F9'].value = '4'
            sheet['G9'].value = '4'
            sheet['H9'].value = '5'
            sheet['I9'].value = '5'
            sheet['J9'].value = '6'
            
            # Revision headers
            sheet.merge_cells('K8:L8')
            sheet['K8'].value = 'Revision - 0'
            sheet.merge_cells('M8:N8')
            sheet['M8'].value = 'Revision - 1'

            for cell in ['A9', 'B9', 'D9', 'F9', 'G9', 'H9', 'I9', 'J9', 'K9', 'M9']:
                sheet[cell].font = bold_font
                sheet[cell].alignment = center_alignment
                sheet[cell].border = border
            sheet.row_dimensions[9].height = 20

            # Row 10: Main column headers
            main_headers = [
                ('A10', 'Sr. No.'),
                ('B10', 'Discipline'),
                ('C10', 'Block'),
                ('D10', ''),
                ('E10', 'Drawing / Document Number'),
                ('F10', 'Name of the Drawing / Document'),
                ('G10', 'Document\nCategories:'),
                ('H10', 'Type - Approval /\nInformation'),
                ('I10', 'Approval\nStatus'),
                ('J10', 'Remarks'),
                ('K10', 'Submission Date'),
                ('L10', 'Reply Date'),
                ('M10', 'Submission Date'),
                ('N10', 'Reply Date'),
            ]

            for cell_ref, text in main_headers:
                cell = sheet[cell_ref]
                cell.value = text
                cell.font = bold_font
                cell.alignment = center_alignment
                cell.border = border
            sheet.row_dimensions[10].height = 35

            # Sample data rows
            sample_data = [
                {
                    "section": "INPUTS",
                    "color": blue_fill,
                    "rows": [
                        {"sr": 1, "disc": "CIV", "block": "P1", "num": "100", "doc_num": "CIL-300 MW-KPI-C-DWG-100", 
                         "name": "Topography survey/Contour Map", "cat": "Drawing", "type": "INFORMATION", 
                         "status": "A", "remarks": "", "sub1": "17-03-2025", "reply1": "", "sub2": "20-03-2025", "reply2": "20-03-2025"},
                        {"sr": 2, "disc": "CIV", "block": "P1", "num": "101", "doc_num": "CIL-300 MW-KPI-C-DWG-101", 
                         "name": "Boundary Layout plan and Co-ordinates", "cat": "Drawing", "type": "APPROVAL", 
                         "status": "A", "remarks": "", "sub1": "01-02-2025", "reply1": "11-02-2025", "sub2": "12-02-2025", "reply2": "20-02-2025"},
                    ]
                },
                {
                    "section": "SLD",
                    "color": blue_fill,
                    "rows": [
                        {"sr": 1, "disc": "ELE", "block": "P1", "num": "200", "doc_num": "CIL-300 MW-KPI-E-DWG-200", 
                         "name": "SLD- AC SYSTEM", "cat": "Drawing", "type": "APPROVAL", 
                         "status": "A", "remarks": "Comments received 22th March, CRS shared with compliance on 25th March", 
                         "sub1": "17-12-2024", "reply1": "", "sub2": "10-03-2025", "reply2": "17-03-2025"},
                        {"sr": 2, "disc": "ELE", "block": "P1", "num": "201", "doc_num": "CIL-300 MW-KPI-E-DWG-201", 
                         "name": "SLD- DC SYSTEM", "cat": "Drawing", "type": "APPROVAL", 
                         "status": "C", "remarks": "Comments received 25th March, compliance submitted on 5-Apr", 
                         "sub1": "20-03-2025", "reply1": "25-03-2025", "sub2": "", "reply2": ""},
                    ]
                },
                {
                    "section": "MECHANICAL STRUCTURE",
                    "color": blue_fill,
                    "rows": [
                        {"sr": 1, "disc": "MEC", "block": "P1", "num": "300", "doc_num": "CIL-300 MW-KPI-M-DOC-300", 
                         "name": "Module Mounting Structure - Design Analysis Report", "cat": "Document", "type": "APPROVAL", 
                         "status": "A", "remarks": "", "sub1": "05-03-2025", "reply1": "13-03-2025", "sub2": "", "reply2": ""},
                    ]
                }
            ]

            current_row = 11
            for section in sample_data:
                # Section header row
                sheet.merge_cells(f'B{current_row}:Q{current_row}')
                cell = sheet[f'B{current_row}']
                cell.value = section["section"]
                cell.font = white_bold_font
                cell.fill = section["color"]
                cell.alignment = center_alignment
                cell.border = border
                
                cell = sheet[f'A{current_row}']
                cell.value = len(sample_data) - len([s for s in sample_data if s == section]) + 1
                cell.font = bold_font
                cell.fill = section["color"]
                cell.alignment = center_alignment
                cell.border = border
                
                sheet.row_dimensions[current_row].height = 25
                current_row += 1

                # Data rows for this section
                for row_data in section["rows"]:
                    cells_data = [
                        (1, row_data["sr"]),
                        (2, row_data["disc"]),
                        (3, row_data["block"]),
                        (4, row_data["num"]),
                        (5, row_data["doc_num"]),
                        (6, row_data["name"]),
                        (7, row_data["cat"]),
                        (8, row_data["type"]),
                        (9, row_data["status"]),
                        (10, row_data["remarks"]),
                        (11, row_data["sub1"]),
                        (12, row_data["reply1"]),
                        (13, row_data["sub2"]),
                        (14, row_data["reply2"]),
                    ]

                    for col, value in cells_data:
                        cell = sheet.cell(row=current_row, column=col)
                        cell.value = value
                        cell.alignment = center_alignment if col <= 9 or col >= 11 else left_alignment
                        cell.border = border
                        
                        # Status color coding
                        if col == 9:
                            if value == "A":
                                cell.fill = green_fill
                            elif value == "C":
                                cell.fill = red_fill
                    
                    sheet.row_dimensions[current_row].height = 30 if row_data["remarks"] else 25
                    current_row += 1

            # Freeze panes
            sheet.freeze_panes = 'B11'

            # Define the file path to save the report
            reports_dir = os.path.join(settings.MEDIA_ROOT, "reports/")
            os.makedirs(reports_dir, exist_ok=True)  # Ensure the directory exists
            datetime_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = os.path.join(reports_dir, f"design_mdl_report_{datetime_str}.xlsx")

            # Save the workbook to the file path
            workbook.save(file_path)

            # Construct the file URL
            file_url = request.build_absolute_uri(settings.MEDIA_URL + f"reports/design_mdl_report_{datetime_str}.xlsx")
            return Response({"status": True, "message": "Design MDL report generated successfully", "file_url": file_url}, status=200)


        except Exception as e:
            logger.error(f"Error generating Design MDL report: {e}", exc_info=True)
            return Response({"error": str(e)}, status=500)